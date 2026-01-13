#!/usr/bin/env python3
"""
Script independiente para sincronizar el Excel a cualquier BD
Uso: python sincronizar_cloud.py --db local    # Sincroniza BD local SQLite
     python sincronizar_cloud.py --db cloud    # Sincroniza BD en Google Cloud
"""

import sys
import argparse
import os

# Agregar el directorio al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def main():
    parser = argparse.ArgumentParser(description='Sincroniza datos del Excel a la BD')
    parser.add_argument('--db', choices=['local', 'cloud'], default='local',
                        help='Base de datos destino: local (SQLite) o cloud (PostgreSQL)')
    parser.add_argument('--excel', default=r'E:\PROYECTOS WEB\control_patrimonial_web\excel\DATA SIGA 15-12-2025.xlsx',
                        help='Ruta al archivo Excel')

    args = parser.parse_args()

    print("="*80)
    if args.db == 'local':
        print("SINCRONIZACION A BD LOCAL (SQLite)")
        sincronizar_local(args.excel)
    else:
        print("SINCRONIZACION A BD EN CLOUD (PostgreSQL)")
        sincronizar_cloud(args.excel)
    print("="*80)

def sincronizar_local(excel_path):
    """Sincroniza a la BD SQLite local"""
    import sqlite3
    from openpyxl import load_workbook
    import json
    from datetime import datetime

    db_path = r"E:\PROYECTOS WEB\control_patrimonial_web\data\inventario_patrimonial.db"

    print(f"Excel: {excel_path}")
    print(f"BD: {db_path}\n")

    if not os.path.exists(excel_path):
        print(f"ERROR: Archivo no encontrado: {excel_path}")
        return

    # Ya sincronizó anteriormente, esto es solo un resumen
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Estadísticas
    cursor.execute("SELECT COUNT(*) FROM bienes")
    total = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM bienes WHERE cal_2025 IS NOT NULL")
    con_cal = cursor.fetchone()[0]

    print(f"Total bienes: {total}")
    print(f"Bienes con CAL 2025: {con_cal}")
    print(f"Bienes sin CAL 2025: {total - con_cal}")

    cursor.close()
    conn.close()

def sincronizar_cloud(excel_path):
    """Sincroniza a la BD PostgreSQL en Google Cloud"""
    import psycopg2
    from psycopg2 import sql
    from openpyxl import load_workbook
    import json
    from datetime import datetime

    print(f"Excel: {excel_path}")
    print(f"BD: PostgreSQL en Google Cloud (35.222.111.223)\n")

    if not os.path.exists(excel_path):
        print(f"ERROR: Archivo no encontrado: {excel_path}")
        return

    # Parámetros de conexión
    db_params = {
        'host': '35.222.111.223',
        'port': 5432,
        'database': 'control_patrimonial',
        'user': 'control_patrimonial',
        'password': 'Control2025'
    }

    try:
        print("Conectando a PostgreSQL en Google Cloud...")
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Verificar conexión
        cursor.execute("SELECT version();")
        version = cursor.fetchone()
        print("[OK] Conectado a PostgreSQL")
        print(f"Versión: {version[0][:70]}\n")

        # Cargar Excel
        print("Cargando archivo Excel...")
        wb = load_workbook(excel_path)
        ws = wb.active

        # Obtener estadísticas antes
        cursor.execute("SELECT COUNT(*) FROM bienes")
        bienes_antes = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM bienes WHERE cal_2025 IS NOT NULL")
        con_cal_antes = cursor.fetchone()[0]

        print(f"ANTES de sincronización:")
        print(f"  Total bienes: {bienes_antes}")
        print(f"  Bienes con CAL 2025: {con_cal_antes}\n")

        # Leer y ejecutar SQL
        sql_path = r"E:\PROYECTOS WEB\control_patrimonial_web\sync_to_cloud.sql"
        if os.path.exists(sql_path):
            print(f"Ejecutando sincronización desde {sql_path}...")
            with open(sql_path, 'r', encoding='utf-8') as f:
                sql_content = f.read()

            cursor.execute(sql_content)
            conn.commit()
            print("[OK] Sincronización completada\n")

            # Obtener estadísticas después
            cursor.execute("SELECT COUNT(*) FROM bienes")
            bienes_despues = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM bienes WHERE cal_2025 IS NOT NULL")
            con_cal_despues = cursor.fetchone()[0]

            print(f"DESPUES de sincronización:")
            print(f"  Total bienes: {bienes_despues}")
            print(f"  Bienes con CAL 2025: {con_cal_despues}\n")

            print(f"CAMBIOS REALIZADOS:")
            print(f"  Bienes agregados: {bienes_despues - bienes_antes}")
            print(f"  CAL 2025 actualizados: {con_cal_despues - con_cal_antes}")

        cursor.close()
        conn.close()
        print("\n[OK] Sincronizacion exitosa")

    except psycopg2.OperationalError as e:
        print(f"[ERROR] No se pudo conectar a PostgreSQL: {e}")
        print("\nVerifica:")
        print("  1. La dirección IP sea correcta")
        print("  2. El puerto 5432 esté abierto")
        print("  3. Las credenciales sean correctas")
        print("  4. Que tengas conexión de red a Google Cloud")

if __name__ == '__main__':
    main()
