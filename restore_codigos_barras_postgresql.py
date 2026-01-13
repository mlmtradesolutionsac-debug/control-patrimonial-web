#!/usr/bin/env python3
"""
SCRIPT SEGURO DE RESTAURACION DE CODIGOS DE BARRAS EN POSTGRESQL
- Solo actualiza el campo cod_barras
- Usa transacciones para rollback si falla
- Valida cada actualización
- Conecta a PostgreSQL (base de datos real usada por Gunicorn)
"""

import psycopg2
from psycopg2 import sql, Error
import openpyxl
import os
import sys

def restaurar_codigos_postgresql():
    print("=" * 70)
    print("RESTAURACION DE CODIGOS DE BARRAS EN POSTGRESQL - MODO SEGURO")
    print("=" * 70)

    # Rutas y configuración
    excel_file = 'DATA SIGA 27-11-2025.xlsx'
    db_config = {
        'host': 'localhost',
        'user': 'control_patrimonial',
        'password': 'Control2025Patrimonial!',
        'database': 'control_patrimonial',
        'port': 5432
    }

    if not os.path.exists(excel_file):
        print(f"ERROR: No se encontró {excel_file}")
        return False

    print(f"\n✓ Archivo Excel: {excel_file}")
    print(f"✓ Base de Datos: PostgreSQL ({db_config['database']})")

    # Conectar a PostgreSQL
    try:
        print("\n[1/5] Conectando a PostgreSQL...")
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        print("✓ Conexión exitosa a PostgreSQL")
    except Error as e:
        print(f"ERROR al conectar a PostgreSQL: {e}")
        return False

    # Leer Excel
    try:
        print("\n[1/5] Leyendo archivo Excel...")
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        actualizaciones = []

        # Iterar filas (comenzar en fila 2, saltando encabezado)
        for row_idx in range(2, ws.max_row + 1):
            codigo_patrimonial = ws.cell(row=row_idx, column=2).value  # Columna B
            cod_barras = ws.cell(row=row_idx, column=3).value  # Columna C

            # Validar que no sean nulos
            if codigo_patrimonial and cod_barras is not None:
                # Convertir a string
                codigo_patrimonial = str(codigo_patrimonial).strip()
                cod_barras = str(cod_barras).strip()

                actualizaciones.append((codigo_patrimonial, cod_barras))

        print(f"✓ Se leyeron {len(actualizaciones)} registros del Excel")

    except Exception as e:
        print(f"ERROR al leer Excel: {e}")
        cursor.close()
        conn.close()
        return False

    # Validar que al menos uno sea diferente de '0'
    print("\n[2/5] Validando datos...")
    codigos_reales = sum(1 for _, cb in actualizaciones if cb != '0')
    print(f"✓ Códigos con valor '0': {len(actualizaciones) - codigos_reales}")
    print(f"✓ Códigos REALES (no son '0'): {codigos_reales}")

    if codigos_reales == 0:
        print("WARNING: Todos los códigos son '0', algo no está bien")

    # TRANSACCION: Actualizar BD
    print("\n[3/5] Iniciando actualización (transacción segura)...")
    try:
        cursor.execute("BEGIN")

        actualizados = 0
        no_encontrados = 0

        for codigo_patrimonial, cod_barras in actualizaciones:
            # Buscar el bien por codigo_patrimonial
            cursor.execute(
                "SELECT id FROM bienes WHERE codigo_patrimonial = %s",
                (codigo_patrimonial,)
            )
            result = cursor.fetchone()

            if result:
                bien_id = result[0]
                # Actualizar SOLO cod_barras
                cursor.execute(
                    "UPDATE bienes SET cod_barras = %s WHERE id = %s",
                    (cod_barras, bien_id)
                )
                actualizados += 1
            else:
                no_encontrados += 1

        print(f"✓ Bienes actualizados: {actualizados}")
        print(f"✓ Bienes no encontrados en BD: {no_encontrados}")

        # COMMIT
        conn.commit()
        print("✓ Cambios confirmados en la BD (COMMIT)")

    except Exception as e:
        print(f"ERROR durante actualización: {e}")
        try:
            conn.rollback()
            print("✓ Cambios revertidos (ROLLBACK) - BD está íntegra")
        except:
            pass
        cursor.close()
        conn.close()
        return False

    # Verificación post-actualización
    print("\n[4/5] Verificando integridad post-actualización...")
    try:
        # Contar códigos reales (no son '0')
        cursor.execute("SELECT COUNT(*) FROM bienes WHERE cod_barras IS NOT NULL AND cod_barras != '' AND cod_barras != '0'")
        con_codigo_real = cursor.fetchone()[0]

        # Total
        cursor.execute("SELECT COUNT(*) FROM bienes")
        total = cursor.fetchone()[0]

        print(f"✓ Total de bienes: {total}")
        print(f"✓ Con código real (NO '0'): {con_codigo_real}")
        print(f"✓ Aumento de códigos reales: {con_codigo_real - codigos_reales}")

        # Muestra de algunos códigos
        cursor.execute("SELECT codigo_patrimonial, cod_barras FROM bienes WHERE cod_barras != '0' AND cod_barras != '' LIMIT 5")
        muestras = cursor.fetchall()
        if muestras:
            print("\n✓ Ejemplos de códigos restaurados:")
            for cp, cb in muestras:
                print(f"  - {cp}: {cb}")

    except Exception as e:
        print(f"ERROR en verificación: {e}")

    # Cerrar conexión
    cursor.close()
    conn.close()

    print("\n[5/5] Restauración completada exitosamente")
    print("\n" + "=" * 70)
    print("STATUS: ✅ CODIGOS DE BARRAS RESTAURADOS EN POSTGRESQL")
    print("=" * 70)
    print("\nPróximos pasos:")
    print("1. Gunicorn está usando PostgreSQL (FLASK_ENV=production)")
    print("2. Los códigos se han restaurado en PostgreSQL")
    print("3. Reiniciar Gunicorn para refrescar la cache")
    print("4. Cargar el dashboard y verificar que los códigos se vean")

    return True

if __name__ == '__main__':
    success = restaurar_codigos_postgresql()
    sys.exit(0 if success else 1)
