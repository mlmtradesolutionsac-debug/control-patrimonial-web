#!/usr/bin/env python3
"""
Script para sincronizar datos del Excel a la base de datos PostgreSQL en Google Cloud
- Inserta bienes nuevos que no están en la BD
- Actualiza CAL 2025 para bienes existentes
"""

import subprocess
import json
from datetime import datetime
from openpyxl import load_workbook
import os

# Configuración
EXCEL_PATH = r"E:\PROYECTOS WEB\control_patrimonial_web\excel\DATA SIGA 15-12-2025.xlsx"
GCP_PROJECT = "steam-outlet-480502-d7"
CLOUDSQL_INSTANCE = "patrimonio-sql-v1"
DB_USER = "control_patrimonial"
DB_NAME = "control_patrimonial"

def run_gcloud_sql(sql_query):
    """Ejecuta una query SQL en la BD PostgreSQL en Google Cloud"""
    cmd = [
        'gcloud', 'sql', 'connect', CLOUDSQL_INSTANCE,
        f'--project={GCP_PROJECT}',
        '--user=postgres'
    ]

    try:
        process = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        stdout, stderr = process.communicate(input=sql_query, timeout=30)

        if process.returncode != 0:
            raise Exception(f"Error ejecutando query: {stderr}")

        return stdout
    except subprocess.TimeoutExpired:
        process.kill()
        raise Exception("Timeout al ejecutar query")

def map_estado(estado_str):
    """Convierte estado de texto a código (b=Bueno, r=Regular, m=Malo)"""
    if not estado_str:
        return None
    estado = estado_str.strip().upper()
    mapping = {
        'BUENO': 'b',
        'B': 'b',
        'REGULAR': 'r',
        'R': 'r',
        'MALO': 'm',
        'M': 'm'
    }
    return mapping.get(estado, None)

def get_sede_id(sede_nombre):
    """Obtiene el ID de sede por nombre"""
    query = f"""
    SELECT id FROM sedes WHERE nombre = '{sede_nombre}' LIMIT 1;
    """
    result = run_gcloud_sql(query).strip()
    return int(result.split('\n')[-1]) if result else None

def get_unidad_id(unidad_nombre):
    """Obtiene el ID de unidad por nombre"""
    query = f"""
    SELECT id FROM unidades WHERE nombre = '{unidad_nombre}' LIMIT 1;
    """
    result = run_gcloud_sql(query).strip()
    return int(result.split('\n')[-1]) if result else None

def insert_bien(bien_data, usuario):
    """Inserta un bien nuevo en la BD"""
    try:
        # Preparar datos
        codigo_patrimonial = bien_data['Codigo Patrimonial']
        cod_barras = bien_data['Codigo de barras']
        denominacion = bien_data['Denominacion'].replace("'", "''")
        marca = bien_data.get('Marca', '').replace("'", "''") if bien_data.get('Marca') else 'NULL'
        serie = bien_data.get('Serie', '').replace("'", "''") if bien_data.get('Serie') else 'NULL'
        modelo = bien_data.get('modelo', '').replace("'", "''") if bien_data.get('modelo') else 'NULL'
        responsable = bien_data.get('Responsable', '').replace("'", "''") if bien_data.get('Responsable') else 'NULL'
        usuario_registro = bien_data.get('Usuario', usuario).replace("'", "''")
        sede_nombre = bien_data.get('Sede', '')
        unidad_nombre = bien_data.get('Dependencia', '')
        estado_str = bien_data.get('Estado')
        estado = map_estado(estado_str)
        observaciones = bien_data.get('observaciones', '')
        if observaciones:
            observaciones = f"'{observaciones.replace(chr(39), chr(39)*2)}'"
        else:
            observaciones = 'NULL'
        cal_2025 = bien_data.get('CAL 2025')

        # Obtener IDs
        sede_id = get_sede_id(sede_nombre)
        unidad_id = get_unidad_id(unidad_nombre)

        if not sede_id:
            print(f"  ERROR: No se encontró sede '{sede_nombre}'")
            return False

        # Construir query INSERT
        marca_sql = f"'{marca}'" if marca != 'NULL' else 'NULL'
        serie_sql = f"'{serie}'" if serie != 'NULL' else 'NULL'
        modelo_sql = f"'{modelo}'" if modelo != 'NULL' else 'NULL'
        responsable_sql = f"'{responsable}'" if responsable != 'NULL' else 'NULL'
        unidad_sql = f"{unidad_id}" if unidad_id else 'NULL'
        estado_sql = f"'{estado}'" if estado else 'NULL'
        cal_2025_sql = f"'{cal_2025}'" if cal_2025 else 'NULL'

        query = f"""
        INSERT INTO bienes (
            codigo_patrimonial, cod_barras, denominacion, marca, serie, modelo,
            responsable, usuario_registro, sede_id, unidad_id, estado,
            observaciones, verificado, cal_2025, fecha_registro
        ) VALUES (
            '{codigo_patrimonial}', '{cod_barras}', '{denominacion}', {marca_sql}, {serie_sql}, {modelo_sql},
            {responsable_sql}, '{usuario_registro}', {sede_id}, {unidad_sql}, {estado_sql},
            {observaciones}, 1, {cal_2025_sql}, NOW()
        ) RETURNING id;
        """

        result = run_gcloud_sql(query).strip()
        bien_id = int(result.split('\n')[-1]) if result else None

        if bien_id:
            # Registrar en historial
            detalle = json.dumps({
                'codigo_patrimonial': codigo_patrimonial,
                'denominacion': denominacion,
                'cal_2025': cal_2025
            }).replace("'", "''")

            hist_query = f"""
            INSERT INTO historial (bien_id, usuario, accion, detalle_json, fecha)
            VALUES ({bien_id}, '{usuario_registro}', 'CREAR', '{detalle}', NOW());
            """

            run_gcloud_sql(hist_query)
            print(f"  [OK] Bien insertado: {codigo_patrimonial} (ID: {bien_id})")
            return True
        else:
            print(f"  ERROR: No se pudo obtener ID del bien insertado")
            return False

    except Exception as e:
        print(f"  ERROR: {e}")
        return False

def update_cal_2025(codigos_a_actualizar, usuario):
    """Actualiza CAL 2025 para múltiples bienes"""
    actualizados = 0
    errores = 0

    for codigo, nuevo_cal in codigos_a_actualizar:
        try:
            # Construir query UPDATE
            query = f"""
            UPDATE bienes SET cal_2025 = '{nuevo_cal}' WHERE codigo_patrimonial = '{codigo}';
            """

            run_gcloud_sql(query)
            actualizados += 1

        except Exception as e:
            print(f"  ERROR al actualizar {codigo}: {e}")
            errores += 1

    return actualizados, errores

def main():
    print("="*80)
    print("SINCRONIZACION EXCEL -> POSTGRESQL EN GOOGLE CLOUD")
    print("="*80)

    # Cargar Excel
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Leer datos del Excel
    excel_bienes = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        if row[1]:  # si hay código patrimonial
            codigo = row[1]
            bien = {
                'N': row[0],
                'Codigo Patrimonial': row[1],
                'Codigo de barras': row[2],
                'Denominacion': row[3],
                'Marca': row[4],
                'Serie': row[5],
                'modelo': row[6],
                'Responsable': row[7],
                'Usuario': row[8],
                'Sede': row[9],
                'Dependencia': row[10],
                'Estado': row[11],
                'observaciones': row[12],
                'Verificado': row[13],
                'SIGA TIPO': row[14],
                'Siga UBICACION': row[15],
                'CAL 2025': row[21]
            }
            excel_bienes[codigo] = bien

    print(f"\nTotal bienes en Excel: {len(excel_bienes)}")

    # Obtener códigos patrimoniales de BD en la nube
    try:
        print("Conectando a BD en Google Cloud...")
        query = "SELECT codigo_patrimonial, cal_2025 FROM bienes;"
        result = run_gcloud_sql(query)

        db_bienes = {}
        for line in result.split('\n')[2:]:  # Saltar encabezados
            if line.strip():
                parts = line.split('|')
                if len(parts) >= 2:
                    codigo = parts[0].strip()
                    cal = parts[1].strip() if parts[1].strip() else None
                    db_bienes[codigo] = cal

        print(f"Total bienes en BD Cloud: {len(db_bienes)}\n")

    except Exception as e:
        print(f"ERROR: No se pudo conectar a BD en la nube: {e}")
        return

    # Identificar bienes nuevos
    bienes_nuevos = [cod for cod in excel_bienes if cod not in db_bienes]
    print(f"BIENES NUEVOS A INSERTAR: {len(bienes_nuevos)}")
    print("="*80)

    if bienes_nuevos:
        for codigo in bienes_nuevos:
            bien = excel_bienes[codigo]
            print(f"Insertando: {codigo} - {bien['Denominacion'][:50]}")
            insert_bien(bien, "ADMIN_SYNC")

    # Identificar bienes a actualizar CAL 2025
    bienes_actualizar = []
    for cod, bien_excel in excel_bienes.items():
        if cod in db_bienes:
            cal_excel = bien_excel['CAL 2025']
            cal_db = db_bienes[cod]
            if cal_excel and cal_db != str(cal_excel):
                bienes_actualizar.append((cod, str(cal_excel)))

    print(f"\nBIENES A ACTUALIZAR CAL 2025: {len(bienes_actualizar)}")
    print("="*80)

    if bienes_actualizar:
        print("Actualizando CAL 2025...")
        actualizados, errores = update_cal_2025(bienes_actualizar, "ADMIN_SYNC")
        print(f"[OK] Actualizados: {actualizados}")
        if errores:
            print(f"[ERROR] Errores: {errores}")

    # Resumen final
    print("\n" + "="*80)
    print("SINCRONIZACION COMPLETADA")
    print("="*80)
    print("[OK] Cambios sincronizados a la nube exitosamente")

if __name__ == '__main__':
    main()
