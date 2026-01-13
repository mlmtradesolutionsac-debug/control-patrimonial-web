#!/usr/bin/env python3
"""
Script para crear un endpoint de sincronización
Debe ejecutarse en el servidor de Cloud Run o local
"""

import os
import sys

# Agregar el directorio de la aplicación al path
sys.path.insert(0, r'E:\PROYECTOS WEB\control_patrimonial_web')

# Configurar variable de ambiente para usar la BD en producción
os.environ['FLASK_ENV'] = 'production'

# Importar la aplicación Flask
try:
    from app import create_app, db
    from app.models_sqlalchemy import Bien
    from openpyxl import load_workbook
    import json
    from datetime import datetime

    # Crear la aplicación
    app = create_app()

    # Hacer el sync dentro del contexto de la aplicación
    with app.app_context():
        print("="*80)
        print("SINCRONIZACION DESDE ENDPOINT")
        print("="*80)

        # Cargar Excel
        excel_path = r"E:\PROYECTOS WEB\control_patrimonial_web\excel\DATA SIGA 15-12-2025.xlsx"
        wb = load_workbook(excel_path)
        ws = wb.active

        # Leer datos del Excel
        excel_bienes = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if row[1]:
                codigo = row[1]
                excel_bienes[codigo] = {
                    'Codigo Patrimonial': row[1],
                    'Codigo de barras': row[2],
                    'Denominacion': row[3],
                    'Marca': row[4],
                    'Serie': row[5],
                    'modelo': row[6],
                    'Responsable': row[7],
                    'Usuario': row[8],
                    'Sede': row[9],
                    'Dependencia': row[10],
                    'Estado': row[11],
                    'observaciones': row[12],
                    'Verificado': row[13],
                    'CAL 2025': row[21]
                }

        print(f"Total bienes en Excel: {len(excel_bienes)}")

        # Obtener bienes de la BD
        db_bienes = db.session.query(Bien.codigo_patrimonial, Bien.cal_2025).all()
        db_bienes_dict = {b[0]: b[1] for b in db_bienes}

        print(f"Total bienes en DB: {len(db_bienes_dict)}")

        # Actualizar CAL 2025
        actualizados = 0
        for codigo, bien_excel in excel_bienes.items():
            if codigo in db_bienes_dict:
                cal_excel = bien_excel['CAL 2025']
                cal_db = db_bienes_dict[codigo]
                if cal_excel and cal_db != cal_excel:
                    # Actualizar
                    bien = db.session.query(Bien).filter_by(codigo_patrimonial=codigo).first()
                    if bien:
                        bien.cal_2025 = cal_excel
                        db.session.commit()
                        actualizados += 1

        print(f"Bienes actualizados: {actualizados}")
        print("[OK] Sincronizacion completada")

except Exception as e:
    print(f"[ERROR] {e}")
    import traceback
    traceback.print_exc()
