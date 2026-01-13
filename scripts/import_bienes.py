#!/usr/bin/env python3
"""
Importador universal de bienes desde Excel

Funciona automáticamente en Desarrollo (SQLite) y Producción (PostgreSQL)

Características:
- Lee archivos XLSX/XLS
- Detecta cambios automáticamente
- Trunca campos según límites de BD
- Crea backup antes de actualizar
- Soporta inserción de nuevos bienes

Uso:
    python scripts/import_bienes.py <ruta/al/archivo.xlsx> [--update] [--insert] [--all]

Opciones:
    --update    Solo actualizar bienes existentes
    --insert    Solo insertar bienes nuevos
    --all       Insertar y actualizar (por defecto)
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import shutil

# Agregar directorio raíz al path
sys.path.insert(0, str(Path(__file__).parent.parent))

from wsgi import app
from app import db
from app.models_sqlalchemy import Bien, Sede, Unidad
from scripts.db_helper import (
    print_db_info, require_confirmation, print_success,
    print_error, print_warning, print_info, print_separator,
    get_db_info
)

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print_error("openpyxl no está instalado")
    print_info("Instala con: pip install openpyxl")
    sys.exit(1)


# Límites de campos según BD
FIELD_LIMITS = {
    'codigo_patrimonial': 50,
    'cod_barras': 100,
    'denominacion': 200,
    'marca': 100,
    'modelo': 100,
    'serie': 100,
    'estado': 1,
    'responsable': 200,
    'usuario_registro': 50,
}


def truncate_field(field_name, value):
    """Trunca valores que exceden el límite del campo."""
    if not value or value == '':
        return value

    limit = FIELD_LIMITS.get(field_name)
    if limit:
        value_str = str(value)
        if len(value_str) > limit:
            print_warning(f"  Truncando {field_name}: '{value_str[:50]}...' -> '{value_str[:limit]}'")
            return value_str[:limit]

    return value


def read_excel(file_path):
    """Lee archivo Excel y retorna lista de diccionarios."""
    file_path = Path(file_path)

    if not file_path.exists():
        print_error(f"Archivo no existe: {file_path}")
        return None

    if not file_path.suffix.lower() in ['.xlsx', '.xls']:
        print_error(f"Archivo debe ser .xlsx o .xls, no {file_path.suffix}")
        return None

    try:
        print_info(f"Leyendo archivo: {file_path}")
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        headers = []
        data = []

        for idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if idx == 1:
                # Primera fila = headers
                headers = [h.lower().strip() if h else f'col_{i}' for i, h in enumerate(row)]
            else:
                # Convertir a diccionario
                row_dict = dict(zip(headers, row))
                if any(row_dict.values()):  # Solo si la fila no está vacía
                    data.append(row_dict)

        print_success(f"Leído {len(data)} registros de Excel")
        return data

    except Exception as e:
        print_error(f"Error leyendo Excel: {str(e)}")
        return None


def backup_sqlite():
    """Crea backup de SQLite antes de actualizar."""
    info = get_db_info()

    if not info['is_sqlite']:
        return None

    db_path = Path(info['database'])
    if not db_path.exists():
        return None

    backup_path = db_path.parent / f"{db_path.name}.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    try:
        shutil.copy2(db_path, backup_path)
        print_success(f"Backup creado: {backup_path}")
        return backup_path
    except Exception as e:
        print_error(f"Error creando backup: {e}")
        return None


def import_bienes(excel_data, modo='all'):
    """
    Importa/actualiza bienes desde datos de Excel.

    modo: 'update', 'insert', 'all'
    """
    print_separator("IMPORTANDO BIENES DESDE EXCEL")

    stats = {
        'actualizados': 0,
        'insertados': 0,
        'errores': 0,
        'truncados': 0
    }

    with app.app_context():
        for idx, row_data in enumerate(excel_data, 1):
            try:
                codigo = row_data.get('codigo_patrimonial', row_data.get('código', '')).strip()

                if not codigo:
                    print_warning(f"  Fila {idx}: Sin código patrimonial, saltando")
                    continue

                # Buscar bien existente
                bien_existente = Bien.query.filter_by(codigo_patrimonial=codigo).first()

                if bien_existente and modo in ['update', 'all']:
                    # ACTUALIZAR
                    try:
                        # Truncar campos largos
                        denominacion = truncate_field('denominacion', row_data.get('denominacion', ''))
                        responsable = truncate_field('responsable', row_data.get('responsable', ''))
                        usuario_registro = truncate_field('usuario_registro', row_data.get('usuario_registro', ''))
                        marca = truncate_field('marca', row_data.get('marca', ''))
                        modelo = truncate_field('modelo', row_data.get('modelo', ''))

                        bien_existente.denominacion = denominacion
                        bien_existente.marca = marca
                        bien_existente.modelo = modelo
                        bien_existente.responsable = responsable
                        bien_existente.estado = row_data.get('estado', bien_existente.estado)
                        bien_existente.usuario_registro = usuario_registro

                        # Campos opcionales
                        if 'ubicacion_texto' in row_data:
                            bien_existente.ubicacion_texto = row_data.get('ubicacion_texto')
                        if 'observaciones' in row_data:
                            bien_existente.observaciones = row_data.get('observaciones')
                        if 'cal_2025' in row_data:
                            bien_existente.cal_2025 = row_data.get('cal_2025')

                        stats['actualizados'] += 1

                    except Exception as e:
                        stats['errores'] += 1
                        print_error(f"  Fila {idx} ({codigo}): {str(e)[:60]}")
                        continue

                elif not bien_existente and modo in ['insert', 'all']:
                    # INSERTAR
                    try:
                        sede_id = row_data.get('sede_id', 1)
                        unidad_id = row_data.get('unidad_id', 1)

                        if not sede_id or sede_id <= 0:
                            sede_id = 1
                        if not unidad_id or unidad_id <= 0:
                            unidad_id = 1

                        # Truncar campos largos
                        denominacion = truncate_field('denominacion', row_data.get('denominacion', ''))
                        responsable = truncate_field('responsable', row_data.get('responsable', ''))
                        usuario_registro = truncate_field('usuario_registro', row_data.get('usuario_registro', ''))
                        marca = truncate_field('marca', row_data.get('marca', ''))
                        modelo = truncate_field('modelo', row_data.get('modelo', ''))

                        nuevo_bien = Bien(
                            codigo_patrimonial=codigo,
                            denominacion=denominacion,
                            marca=marca,
                            modelo=modelo,
                            responsable=responsable,
                            usuario_registro=usuario_registro,
                            estado=row_data.get('estado', 'Bueno'),
                            ubicacion_texto=row_data.get('ubicacion_texto', ''),
                            observaciones=row_data.get('observaciones', ''),
                            sede_id=sede_id,
                            unidad_id=unidad_id,
                            fecha_registro=datetime.now()
                        )

                        db.session.add(nuevo_bien)
                        stats['insertados'] += 1

                    except Exception as e:
                        stats['errores'] += 1
                        print_error(f"  Fila {idx} ({codigo}): {str(e)[:60]}")
                        continue

                # Commit cada 100 registros
                if idx % 100 == 0:
                    try:
                        db.session.commit()
                        pct = int(idx / len(excel_data) * 100)
                        print_info(f"  {idx}/{len(excel_data)} ({pct}%) - A: {stats['actualizados']}, I: {stats['insertados']}, E: {stats['errores']}")
                    except Exception as e:
                        db.session.rollback()
                        print_error(f"Error al guardar: {str(e)}")

            except Exception as e:
                stats['errores'] += 1
                print_error(f"  Fila {idx}: Error inesperado - {str(e)[:60]}")

        # Commit final
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print_error(f"Error en commit final: {str(e)}")

    # Resumen
    print_separator("IMPORTACIÓN COMPLETADA")
    print(f"Actualizados: {stats['actualizados']}")
    print(f"Insertados:   {stats['insertados']}")
    print(f"Errores:      {stats['errores']}")

    total = Bien.query.count()
    print(f"\nTotal bienes en BD: {total}")

    return stats


def show_help():
    """Muestra ayuda."""
    help_text = """
IMPORTADOR DE BIENES - Control Patrimonial

Importa/actualiza bienes desde archivos Excel (.xlsx o .xls)
Funciona automáticamente en Desarrollo y Producción

COMANDOS:

  python scripts/import_bienes.py archivo.xlsx
    Importar Excel (inserta nuevos + actualiza existentes)

  python scripts/import_bienes.py archivo.xlsx --update
    Solo actualizar bienes existentes

  python scripts/import_bienes.py archivo.xlsx --insert
    Solo insertar bienes nuevos

FORMATO DEL EXCEL:
La primera fila debe contener encabezados. Campos soportados:

Requeridos:
  codigo_patrimonial    Código único del bien

Opcionales:
  denominacion          Nombre/descripción
  marca                 Marca del equipo
  modelo                Modelo
  serie                 Número de serie
  responsable           Persona responsable
  estado                Estado del bien
  usuario_registro      Usuario que registra
  ubicacion_texto       Ubicación del bien
  observaciones         Notas
  cal_2025              Calibración 2025
  sede_id               ID de sede
  unidad_id             ID de unidad

NOTAS:
- Los campos se truncan automáticamente según límites de BD
- Se crea backup de SQLite antes de importar
- Los cambios se aplican en batches de 100 registros
- Respeta la tabla actual (SQLite o PostgreSQL)
"""
    print(help_text)


def main():
    """Función principal."""
    if len(sys.argv) < 2 or sys.argv[1] in ['help', '--help', '-h']:
        show_help()
        return

    excel_file = sys.argv[1]
    modo = 'all'

    if '--update' in sys.argv:
        modo = 'update'
    elif '--insert' in sys.argv:
        modo = 'insert'

    # Leer Excel
    excel_data = read_excel(excel_file)
    if not excel_data:
        return

    print_db_info()

    # Mostrar preview
    print_separator("PREVIEW - Primeros 3 registros")
    for i, row in enumerate(excel_data[:3], 1):
        print(f"  {i}. {row}")

    # Crear backup si es SQLite
    if get_db_info()['is_sqlite']:
        backup_sqlite()

    # Pedir confirmación si hay muchos cambios
    if len(excel_data) > 100:
        if not require_confirmation(f"¿Importar {len(excel_data)} bienes? Esto puede tomar un tiempo."):
            print_warning("Importación cancelada")
            return

    # Importar
    import_bienes(excel_data, modo)


if __name__ == '__main__':
    main()
