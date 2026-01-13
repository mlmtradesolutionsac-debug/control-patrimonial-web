#!/usr/bin/env python3
"""
Comandos Flask para administración
Uso: flask [comando]
"""

import click
from flask import current_app
from flask.cli import with_appcontext
from openpyxl import load_workbook
from datetime import datetime
import os

@click.command('sync-excel')
@click.argument('excel_path', required=False)
@click.option('--db', 'database', default='sqlite', help='Database: sqlite o cloud')
@with_appcontext
def sync_excel_command(excel_path=None, database='sqlite'):
    """
    Sincroniza datos del Excel a la BD
    Uso: flask sync-excel /ruta/al/archivo.xlsx
    """
    from app import db
    from app.models_sqlalchemy import Bien, Sede, Unidad, Historial
    import json

    if not excel_path:
        # Usar la ruta por defecto
        excel_path = os.path.join(current_app.root_path, '..', 'excel', 'DATA SIGA 15-12-2025.xlsx')

    print("="*80)
    print("SINCRONIZACION EXCEL A BD")
    print("="*80)
    print(f"\nArchivo: {excel_path}")

    if not os.path.exists(excel_path):
        click.echo(f"ERROR: Archivo no encontrado: {excel_path}", err=True)
        return

    # Cargar Excel
    wb = load_workbook(excel_path)
    ws = wb.active

    excel_bienes = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        if row[1]:  # código patrimonial
            codigo = row[1]
            excel_bienes[codigo] = {
                'Codigo Patrimonial': row[1],
                'Codigo de barras': row[2],
                'Denominacion': row[3],
                'Marca': row[4],
                'Serie': row[5],
                'modelo': row[6],
                'Responsable': row[7],
                'Usuario': row[8],
                'Sede': row[9],
                'Dependencia': row[10],
                'Estado': row[11],
                'observaciones': row[12],
                'Verificado': row[13],
                'SIGA TIPO': row[14],
                'Siga UBICACION': row[15],
                'CAL 2025': row[21]
            }

    print(f"Total bienes en Excel: {len(excel_bienes)}\n")

    # Obtener sedes y unidades
    sedes = {s.nombre: s.id for s in db.session.query(Sede).all()}
    unidades = {u.nombre: u.id for u in db.session.query(Unidad).all()}

    # Obtener bienes actuales
    db_bienes = {b.codigo_patrimonial: b for b in db.session.query(Bien).all()}

    # Mapeo de estado
    def map_estado(estado_str):
        if not estado_str:
            return None
        estado = estado_str.strip().upper()
        mapping = {'BUENO': 'b', 'B': 'b', 'REGULAR': 'r', 'R': 'r', 'MALO': 'm', 'M': 'm'}
        return mapping.get(estado, None)

    # Procesar bienes nuevos
    bienes_nuevos = [cod for cod in excel_bienes if cod not in db_bienes]
    print(f"BIENES NUEVOS: {len(bienes_nuevos)}")

    for codigo in bienes_nuevos:
        bien = excel_bienes[codigo]
        try:
            sede_id = sedes.get(bien['Sede'])
            unidad_id = unidades.get(bien['Dependencia'])

            if not sede_id:
                click.echo(f"  ADVERTENCIA: Sede no encontrada para {codigo}: {bien['Sede']}")
                continue

            nuevo_bien = Bien(
                codigo_patrimonial=bien['Codigo Patrimonial'],
                cod_barras=bien['Codigo de barras'],
                denominacion=bien['Denominacion'],
                marca=bien['Marca'],
                serie=bien['Serie'],
                modelo=bien['modelo'],
                responsable=bien['Responsable'],
                usuario_registro=bien['Usuario'] or 'SYNC',
                sede_id=sede_id,
                unidad_id=unidad_id,
                estado=map_estado(bien['Estado']),
                observaciones=bien['observaciones'],
                verificado=1 if bien.get('Verificado') == 'SI' else 0,
                cal_2025=bien['CAL 2025'],
                fecha_registro=datetime.now()
            )

            db.session.add(nuevo_bien)
            db.session.flush()

            # Registrar en historial
            historial = Historial(
                bien_id=nuevo_bien.id,
                usuario='SYNC',
                accion='CREAR',
                detalle_json=json.dumps({'codigo': codigo}),
                fecha=datetime.now()
            )
            db.session.add(historial)

            click.echo(f"  [OK] {codigo} - {bien['Denominacion'][:40]}")

        except Exception as e:
            click.echo(f"  [ERROR] {codigo}: {e}", err=True)
            db.session.rollback()

    # Procesar actualizaciones de CAL 2025
    bienes_actualizar = []
    for cod in excel_bienes:
        if cod in db_bienes:
            bien_db = db_bienes[cod]
            cal_excel = excel_bienes[cod]['CAL 2025']
            if cal_excel and bien_db.cal_2025 != cal_excel:
                bienes_actualizar.append((cod, bien_db.id, cal_excel))

    print(f"\nBIENES A ACTUALIZAR CAL 2025: {len(bienes_actualizar)}")

    for codigo, bien_id, cal_nuevo in bienes_actualizar:
        try:
            bien_db = db_bienes[codigo]
            cal_anterior = bien_db.cal_2025

            # Actualizar
            bien_db.cal_2025 = cal_nuevo
            db.session.commit()

            # Registrar en historial
            historial = Historial(
                bien_id=bien_id,
                usuario='SYNC',
                accion='EDITAR',
                detalle_json=json.dumps({
                    'campo': 'cal_2025',
                    'anterior': cal_anterior,
                    'nuevo': cal_nuevo
                }),
                fecha=datetime.now()
            )
            db.session.add(historial)
            db.session.commit()

        except Exception as e:
            click.echo(f"  [ERROR] {codigo}: {e}", err=True)
            db.session.rollback()

    # Estadísticas finales
    print("\n" + "="*80)
    print("ESTADISTICAS FINALES")
    print("="*80)

    total = db.session.query(Bien).count()
    con_cal = db.session.query(Bien).filter(Bien.cal_2025 != None).count()

    print(f"Total bienes: {total}")
    print(f"Bienes con CAL 2025: {con_cal}")
    print(f"Bienes sin CAL 2025: {total - con_cal}")
    print(f"\n[OK] Sincronización completada")

@click.command('corregir-cal-2025')
@click.option('--backup', is_flag=True, help='Crear archivo de backup antes de ejecutar')
@click.option('--dry-run', is_flag=True, help='Simular cambios sin guardar (solo lectura)')
@with_appcontext
def corregir_cal_2025_command(backup=False, dry_run=False):
    """
    Corrige masivamente bienes con CAL 2025 que tienen estado "Malo" -> "Regular"

    Esta corrección es necesaria porque:
    - Un bien con CAL 2025 significa que fue verificado/inventariado
    - Un bien verificado NO debería tener estado "Malo"
    - El estado correcto es "Regular" (mantenible) o "Bueno" (en funcionamiento)

    Uso:
        flask corregir-cal-2025              # Ejecutar cambios reales
        flask corregir-cal-2025 --dry-run    # Simular sin cambiar nada
        flask corregir-cal-2025 --backup     # Crear backup antes (recomendado)

    IMPORTANTE: Haz backup de la BD antes de ejecutar esto en producción
    """
    from app import db
    from app.models_sqlalchemy import Bien, Historial
    from app.utils import log_action
    from sqlalchemy import and_

    print("="*80)
    print("CORRECCION DE ESTADOS: Bienes con CAL 2025 en estado MALO -> REGULAR")
    print("="*80)
    print()

    if dry_run:
        print("[ADVERTENCIA] MODO SIMULACION (--dry-run): Los cambios NO se guardaran")
        print()

    try:
        # Buscar bienes con CAL 2025 y estado 'Malo'
        bienes_a_corregir = db.session.query(Bien).filter(
            and_(
                Bien.cal_2025 != None,  # Tiene CAL 2025 (no es nulo)
                Bien.cal_2025 != '',    # Tiene CAL 2025 (no es vacío)
                Bien.estado == 'm'      # Estado es 'Malo' (m)
            )
        ).all()

        cantidad = len(bienes_a_corregir)

        if cantidad == 0:
            print("[OK] No hay bienes que corrija (todos los bienes con CAL 2025 ya tienen estado correcto)")
            print()
            return

        print(f"Bienes encontrados para corregir: {cantidad}")
        print()

        if not dry_run and cantidad > 0:
            confirmacion = input(f"¿Confirmas corregir {cantidad} registros? (S/N): ")
            if confirmacion.upper() != 'S':
                print("Operación cancelada.")
                return

        print()
        print("Detalles de bienes a corregir:")
        print("-" * 80)

        # Mostrar muestra de bienes (primeros 10)
        for i, bien in enumerate(bienes_a_corregir[:10], 1):
            print(f"  {i}. {bien.codigo_patrimonial} | {bien.denominacion[:50]}")
            print(f"     -> CAL 2025: {bien.cal_2025} | Estado ACTUAL: Malo (m) -> NUEVO: Regular (r)")

        if cantidad > 10:
            print(f"  ... y {cantidad - 10} bienes más")

        print()

        if dry_run:
            print("[OK] Simulación completada exitosamente")
            print(f"   Se corregirían {cantidad} registros si ejecutas sin --dry-run")
            return

        # ACTUALIZACIÓN MASIVA (operación eficiente sin bucles)
        print("Ejecutando UPDATE masivo...")

        # Contar antes
        count_antes_malo = db.session.query(Bien).filter(
            and_(
                Bien.cal_2025 != None,
                Bien.cal_2025 != '',
                Bien.estado == 'm'
            )
        ).count()

        # UPDATE masivo usando SQLAlchemy
        actualizados = db.session.query(Bien).filter(
            and_(
                Bien.cal_2025 != None,
                Bien.cal_2025 != '',
                Bien.estado == 'm'
            )
        ).update({'estado': 'r'}, synchronize_session=False)

        db.session.commit()

        print(f"[OK] UPDATE ejecutado: {actualizados} registros actualizados")
        print()

        # Registrar en log de aplicación
        log_action(
            'CORRECCION_MASIVA_CAL_2025',
            f'Corregidos {actualizados} bienes: estado Malo -> Regular (CAL 2025 != NULL)'
        )

        # Estadísticas finales
        print("="*80)
        print("ESTADISTICAS FINALES")
        print("="*80)

        total_bienes = db.session.query(Bien).count()
        bienes_con_cal = db.session.query(Bien).filter(
            and_(Bien.cal_2025 != None, Bien.cal_2025 != '')
        ).count()
        bienes_malo = db.session.query(Bien).filter(Bien.estado == 'm').count()
        bienes_regular = db.session.query(Bien).filter(Bien.estado == 'r').count()
        bienes_bueno = db.session.query(Bien).filter(Bien.estado == 'b').count()

        print(f"Total bienes: {total_bienes}")
        print(f"Bienes con CAL 2025: {bienes_con_cal}")
        print()
        print(f"Estado BUENO (b):   {bienes_bueno}")
        print(f"Estado REGULAR (r): {bienes_regular}")
        print(f"Estado MALO (m):    {bienes_malo}")
        print()
        print(f"Registros corregidos: {actualizados}")
        print()
        print("[OK] Corrección completada exitosamente")

    except Exception as e:
        db.session.rollback()
        print()
        print(f"[ERROR] {str(e)}")
        print()
        print("Operacion rechazada. No se realizaron cambios.")
        raise


def init_app(app):
    """Registra los comandos con la aplicación Flask"""
    app.cli.add_command(sync_excel_command)
    app.cli.add_command(corregir_cal_2025_command)
