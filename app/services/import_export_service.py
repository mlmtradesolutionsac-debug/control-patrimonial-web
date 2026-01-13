"""
FASE 2: Servicio de Importación/Exportación de Bienes

Funcionalidades:
- Importar bienes desde Excel/CSV
- Exportar bienes a CSV/Excel/PDF
- Validación por fila durante importación
- Manejo robusto de errores
"""

import io
import csv
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from app.repository import RepoSQLAlchemy as Repo
from app.utils import validate_bien_data


class ImportService:
    """Servicio de importación de bienes desde archivos Excel/CSV"""

    @staticmethod
    def importar_excel(archivo_path, usuario_actual):
        """
        Importar bienes desde archivo Excel

        Args:
            archivo_path: Ruta al archivo Excel
            usuario_actual: Usuario que realiza la importación

        Returns:
            dict con estructura:
            {
                'success': bool,
                'creados': int,
                'errores': list,
                'total_procesados': int,
                'mensaje': str
            }
        """
        try:
            # Leer archivo Excel
            df = pd.read_excel(archivo_path)
            repo = Repo()

            # Cargar mapas de Sede/Unidad (Nombre -> ID)
            sedes_map = {str(s[1]).strip().upper(): str(s[0]) for s in repo.obtener_sedes()}
            unidades_map = {str(u[1]).strip().upper(): str(u[0]) for u in repo.obtener_unidades()}

            errores = []
            creados = 0
            actualizados = 0 

            # Helper para buscar columnas insensible a case/espacios
            def get_col_val(row, col_name_pattern):
                # Buscar columna que coincida
                for col in df.columns:
                    if str(col).strip().upper() == col_name_pattern.upper():
                        val = row.get(col)
                        # Manejar NaN/None
                        if pd.isna(val): return ''
                        return str(val).strip()
                
                # Fallback a variantes comunes
                if col_name_pattern == 'CODIGO PATRIMONIAL': return get_col_val(row, 'CODIGO')
                if col_name_pattern == 'DENOMINACION': return get_col_val(row, 'DESCRIPCION')
                return ''

            # Helper para normalizar estado
            def normalizar_estado(valor):
                val = str(valor).strip().lower()
                if val in ['b', 'bueno', 'nuevo', 'buen estado', 'operativo']: return 'b'
                if val in ['r', 'regular']: return 'r'
                if val in ['m', 'malo', 'mal estado', 'inoperativo']: return 'm'
                return 'r' # Valor por defecto seguro

            # Procesar cada fila
            for idx, row in df.iterrows():
                try:
                    # Obtener nombres y buscar IDs
                    sede_nombre = get_col_val(row, 'SEDE')
                    unidad_nombre = get_col_val(row, 'DEPENDENCIA')
                    
                    sede_id = sedes_map.get(sede_nombre.upper(), '')
                    unidad_id = unidades_map.get(unidad_nombre.upper(), '')

                    # Construir diccionario de datos
                    # Ajustado a nombres reales del Excel (sin acentos en algunos casos)
                    data = {
                        'codigo_patrimonial': get_col_val(row, 'CODIGO PATRIMONIAL'),
                        'denominacion': get_col_val(row, 'DENOMINACION'),
                        'sede': sede_id, 
                        'estado': normalizar_estado(get_col_val(row, 'ESTADO')), # Normalizar estado
                        'dependencia': unidad_id, 
                        'marca': get_col_val(row, 'MARCA'),
                        'modelo': get_col_val(row, 'MODELO'),
                        'serie': get_col_val(row, 'SERIE'),
                        'ubicacion': get_col_val(row, 'SIGA UBICACIÓN'), 
                        'descripcion': get_col_val(row, 'OBSERVACIONES'),
                        'codigo_barras': get_col_val(row, 'CODIGO DE BARRAS'),
                        # Campos CAL
                        'cal_2020': get_col_val(row, 'CAL 2020'),
                        'cal_2021': get_col_val(row, 'CAL 2021'),
                        'cal_2022': get_col_val(row, 'CAL 2022'),
                        'cal_2023': get_col_val(row, 'CAL 2023'),
                        'cal_2024': get_col_val(row, 'CAL 2024'),
                        'cal_2025': get_col_val(row, 'CAL 2025'),
                    }

                    # Validar datos
                    validation_errors = validate_bien_data(data)

                    if validation_errors:
                        errores.append({
                            'fila': idx + 2,
                            'codigo': data.get('codigo_patrimonial', 'SIN_CÓDIGO'),
                            'errores': validation_errors
                        })
                    else:
                        # Verificar si el bien existe
                        bien_existente = repo.obtener_bien_por_codigo(data['codigo_patrimonial'])
                        
                        if bien_existente:
                            # Actualizar bien existente
                            repo.actualizar_bien(bien_existente['id'], data, usuario_actual)
                            actualizados += 1
                        else:
                            # Crear bien nuevo
                            repo.crear_bien(data, usuario_actual)
                            creados += 1

                except Exception as e:
                    errores.append({
                        'fila': idx + 2,
                        'error': f'Error procesando fila: {str(e)}'
                    })

            return {
                'success': len(errores) == 0,
                'creados': creados,
                'actualizados': actualizados,
                'errores': errores,
                'total_procesados': len(df),
                'mensaje': f'Importados {creados}, Actualizados {actualizados} de {len(df)} bienes'
            }

        except Exception as e:
            return {
                'success': False,
                'creados': 0,
                'actualizados': 0,
                'errores': [f'Error al leer archivo: {str(e)}'],
                'total_procesados': 0,
                'mensaje': 'Error en la importación'
            }

    @staticmethod
    def importar_csv(archivo_path, usuario_actual):
        """Importar desde CSV (compatible con Excel)"""
        try:
            df = pd.read_csv(archivo_path, encoding='utf-8')
            repo = Repo()
            # Reutilizar lógica de Excel
            return ImportService._procesar_dataframe(df, repo, usuario_actual)
        except UnicodeDecodeError:
            # Reintentar con encoding diferente
            try:
                df = pd.read_csv(archivo_path, encoding='latin-1')
                repo = Repo()
                return ImportService._procesar_dataframe(df, repo, usuario_actual)
            except Exception as e:
                return {
                    'success': False,
                    'creados': 0,
                    'errores': [f'Error al leer CSV: {str(e)}'],
                    'total_procesados': 0,
                    'mensaje': 'Error en la importación'
                }

    @staticmethod
    def _procesar_dataframe(df, repo, usuario_actual):
        """Procesar DataFrame de pandas"""
        errores = []
        creados = 0
        actualizados = 0 # Contador para bienes actualizados

        for idx, row in df.iterrows():
            try:
                data = {
                    'codigo_patrimonial': str(row.get('Código Patrimonial', '')).strip(),
                    'denominacion': str(row.get('Denominación', '')).strip(),
                    'sede': str(row.get('Sede ID', '')).strip(),
                    'estado': str(row.get('Estado', '')).strip(),
                    'unidad': str(row.get('Unidad', '')).strip() if 'Unidad' in row.index else '',
                    'marca': str(row.get('Marca', '')).strip() if 'Marca' in row.index else '',
                    'modelo': str(row.get('Modelo', '')).strip() if 'Modelo' in row.index else '',
                    'serie': str(row.get('Serie', '')).strip() if 'Serie' in row.index else '',
                    'ubicacion': str(row.get('Ubicación', '')).strip() if 'Ubicación' in row.index else '',
                    # Campos CAL
                    'cal_2020': str(row.get('CAL 2020', '')).strip() if 'CAL 2020' in row.index else '',
                    'cal_2021': str(row.get('CAL 2021', '')).strip() if 'CAL 2021' in row.index else '',
                    'cal_2022': str(row.get('CAL 2022', '')).strip() if 'CAL 2022' in row.index else '',
                    'cal_2023': str(row.get('CAL 2023', '')).strip() if 'CAL 2023' in row.index else '',
                    'cal_2024': str(row.get('CAL 2024', '')).strip() if 'CAL 2024' in row.index else '',
                    'cal_2025': str(row.get('CAL 2025', '')).strip() if 'CAL 2025' in row.index else '',
                }

                validation_errors = validate_bien_data(data)

                if validation_errors:
                    errores.append({
                        'fila': idx + 2,
                        'codigo': data.get('codigo_patrimonial', 'SIN_CÓDIGO'),
                        'errores': validation_errors
                    })
                else:
                    bien_existente = repo.obtener_bien_por_codigo(data['codigo_patrimonial'])
                    if bien_existente:
                        repo.actualizar_bien(bien_existente['id'], data, usuario_actual)
                        actualizados += 1
                    else:
                        repo.crear_bien(data, usuario_actual)
                        creados += 1

            except Exception as e:
                errores.append({
                    'fila': idx + 2,
                    'error': f'Error: {str(e)}'
                })

        return {
            'success': len(errores) == 0,
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores,
            'total_procesados': len(df),
            'mensaje': f'Importados {creados}, Actualizados {actualizados} de {len(df)} bienes'
        }


class ExportService:
    """Servicio de exportación de bienes a múltiples formatos"""

    @staticmethod
    def exportar_csv(filtros=None):
        """Exportar bienes a CSV"""
        repo = Repo()
        bienes = repo.listar_bienes(filtros or {}, limit=100000)

        output = io.StringIO()
        if not bienes:
            return output

        # Campos a exportar
        campos = [
            'id', 'codigo_patrimonial', 'denominacion', 'estado',
            'sede_nombre', 'unidad_nombre', 'marca', 'modelo',
            'serie', 'ubicacion', 'fecha_registro'
        ]

        writer = csv.DictWriter(output, fieldnames=campos)
        writer.writeheader()

        for bien in bienes:
            fila = {}
            for campo in campos:
                valor = bien.get(campo, '')
                # Convertir datetime a string
                if hasattr(valor, 'isoformat'):
                    valor = valor.isoformat()
                fila[campo] = valor
            writer.writerow(fila)

        return output.getvalue()

    @staticmethod
    def exportar_excel(filtros=None, titulo='Reporte de Bienes'):
        """Exportar bienes a Excel con formato"""
        repo = Repo()
        bienes = repo.listar_bienes(filtros or {}, limit=100000)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Bienes'

        # Estilos
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        # Encabezados (nuevos, ordenados según lo solicitado)
        headers = [
            'Código Patrimonial', 'Código Barras', 'Denominación', 'Marca', 'Modelo', 'Serie',
            'Responsable', 'Sede', 'Dependencia', 'Estado', 'CAL 2025', 'Ubicación',
            'Inventariador', 'Fecha Registro'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Datos (nuevos, en el mismo orden que los encabezados)
        for row_num, bien in enumerate(bienes, 2):
            ws.cell(row=row_num, column=1, value=bien.get('codigo_patrimonial', ''))
            ws.cell(row=row_num, column=2, value=bien.get('cod_barras', ''))
            ws.cell(row=row_num, column=3, value=bien.get('denominacion', ''))
            ws.cell(row=row_num, column=4, value=bien.get('marca', ''))
            ws.cell(row=row_num, column=5, value=bien.get('modelo', ''))
            ws.cell(row=row_num, column=6, value=bien.get('serie', ''))
            ws.cell(row=row_num, column=7, value=bien.get('responsable', ''))
            ws.cell(row=row_num, column=8, value=bien.get('sede', ''))
            ws.cell(row=row_num, column=9, value=bien.get('dependencia', ''))
            ws.cell(row=row_num, column=10, value=bien.get('estado', ''))

            # CAL 2025: extraer solo el número (remover "CAL-" o "CAL'" si existe)
            cal_2025 = bien.get('cal_2025', '')
            if cal_2025 and isinstance(cal_2025, str):
                if cal_2025.startswith("CAL-"):
                    cal_2025 = cal_2025.replace("CAL-", "", 1).strip()
                elif cal_2025.startswith("CAL'"):
                    cal_2025 = cal_2025.replace("CAL'", "", 1).strip()
            ws.cell(row=row_num, column=11, value=cal_2025)

            ws.cell(row=row_num, column=12, value=bien.get('ubicacion_texto', ''))
            ws.cell(row=row_num, column=13, value=bien.get('usuario_registro', ''))

            # Formatear fecha de manera más legible (sin segundos)
            fecha = bien.get('fecha_registro', '')
            if fecha and isinstance(fecha, str):
                # Si viene en formato ISO (2025-11-27T10:27:20.028195), convertir a formato legible
                try:
                    # Extraer solo fecha y hora sin segundos: 2025-11-27 10:27
                    fecha_sin_ms = fecha.split('.')[0]  # Quita .028195 → 2025-11-27T10:27:20
                    fecha_parte = fecha_sin_ms.split('T')[0]  # Quita hora → 2025-11-27
                    hora_parte = fecha_sin_ms.split('T')[1][:5] if 'T' in fecha_sin_ms else ''  # Primeros 5 chars de hora → 10:27
                    fecha = f"{fecha_parte} {hora_parte}".strip()
                except:
                    pass
            ws.cell(row=row_num, column=14, value=fecha)

        # Auto-ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value or ''))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column].width = adjusted_width

        return wb

    @staticmethod
    def exportar_pdf(filtros=None, titulo='Reporte de Bienes Patrimoniales'):
        """Exportar bienes a PDF"""
        repo = Repo()
        bienes = repo.listar_bienes(filtros or {}, limit=1000)  # Limitar para no hacer PDF muy grande

        # Crear PDF en memoria
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=20, bottomMargin=20)
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=20,
            alignment=1  # Centrado
        )

        # Título
        elements.append(Paragraph(titulo, title_style))
        elements.append(Spacer(1, 0.3*inch))

        # Información de generación
        fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=9)
        elements.append(Paragraph(
            f'Generado: {fecha_gen} | Total de bienes: {len(bienes)}',
            info_style
        ))
        elements.append(Spacer(1, 0.2*inch))

        # Tabla de datos
        data = [['Código', 'Denominación', 'Estado', 'Sede', 'Ubicación']]

        for bien in bienes[:100]:  # Máximo 100 filas por PDF
            data.append([
                str(bien.get('codigo_patrimonial', ''))[:15],
                str(bien.get('denominacion', ''))[:30],
                str(bien.get('estado', ''))[:10],
                str(bien.get('sede_nombre', ''))[:15],
                str(bien.get('ubicacion', ''))[:20]
            ])

        table = Table(data, colWidths=[1.2*inch, 2*inch, 0.8*inch, 1.2*inch, 1.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)

        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer

    @staticmethod
    def generar_plantilla_importacion():
        """Generar plantilla Excel vacía para importación"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bienes'

        # Estilos para header
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Headers
        headers = [
            'Código Patrimonial',
            'Denominación',
            'Sede ID',
            'Estado',
            'Unidad',
            'Marca',
            'Modelo',
            'Serie',
            'Ubicación'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Ejemplo de datos
        ejemplo = [
            'PATRI-2025-001',
            'Computadora HP',
            '1',
            'Bueno',
            'Administración',
            'HP',
            'ProDesk',
            'SN-123456',
            'Oficina 101'
        ]

        for col_num, valor in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col_num, value=valor)

        # Auto-ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value or ''))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width

        # Agregar comentarios/instrucciones
        ws['A1'].comment = None  # Se agregaría un comentario si se necesita

        return wb


# Exportar funciones para uso directo
__all__ = ['ImportService', 'ExportService']
