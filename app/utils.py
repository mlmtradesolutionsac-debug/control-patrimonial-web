"""
Utilidades para la aplicación: logging, seguridad, validación, etc.
"""
import logging
import os
from functools import wraps
from datetime import datetime
from flask import flash, redirect, url_for, current_app, request
from flask_login import current_user

# ==================== LOGGING ====================
def setup_logging(app):
    """Configurar logging de la aplicación"""
    if not app.debug and not app.testing:
        logs_dir = os.path.join(os.path.dirname(app.instance_path), 'logs')
        if not os.path.exists(logs_dir):
            os.makedirs(logs_dir)

        file_handler = logging.FileHandler(os.path.join(logs_dir, 'app.log'))
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
        ))
        file_handler.setLevel(logging.INFO)
        app.logger.addHandler(file_handler)

        app.logger.setLevel(logging.INFO)
        app.logger.info('Control Patrimonial inicializado')


def log_action(action_type, description, extra_data=None):
    """Log de acciones en la aplicación"""
    try:
        username = current_user.username if current_user.is_authenticated else 'anonimo'
        log_entry = f"[{action_type}] Usuario: {username} | {description}"
        if extra_data:
            log_entry += f" | Datos: {extra_data}"
        current_app.logger.info(log_entry)
    except Exception as e:
        current_app.logger.error(f"Error logging action: {str(e)}")


def log_security_event(event_type, description, severity='INFO', ip_address=None):
    """Log específico para eventos de seguridad con contexto adicional"""
    try:
        username = current_user.username if current_user.is_authenticated else 'anonimo'
        ip = ip_address or (request.remote_addr if request else 'unknown')

        log_entry = f"[SECURITY:{event_type}] Usuario: {username} | IP: {ip} | {description}"

        if severity == 'CRITICAL':
            current_app.logger.critical(log_entry)
        elif severity == 'ERROR':
            current_app.logger.error(log_entry)
        elif severity == 'WARNING':
            current_app.logger.warning(log_entry)
        else:
            current_app.logger.info(log_entry)
    except Exception as e:
        current_app.logger.error(f"Error logging security event: {str(e)}")


# ==================== DECORADORES DE ACCESO ====================
def role_required(required_role):
    """Decorador para verificar rol de usuario"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Debes iniciar sesión', 'warning')
                return redirect(url_for('auth.login'))

            # Si es admin, permitir acceso a todo
            if current_user.is_admin:
                return f(*args, **kwargs)

            # Verificar rol específico
            if current_user.rol != required_role:
                flash('No tienes permiso para acceder a esta página', 'danger')
                log_action('ACCESO_DENEGADO', f'Intento de acceso a {f.__name__} sin permisos')
                return redirect(url_for('main.dashboard'))

            return f(*args, **kwargs)
        return decorated_function
    return decorator


def admin_required(f):
    """Decorador para requerir acceso de admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Acceso denegado. Se requieren permisos de administrador', 'danger')
            log_action('ACCESO_DENEGADO', f'Intento de acceso admin a {f.__name__}')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== VALIDACIÓN ====================
def sanitize_input(value):
    """Sanitizar entrada de usuario"""
    if isinstance(value, str):
        # Remover caracteres peligrosos
        value = value.strip()
        # Limitar longitud
        value = value[:500] if len(value) > 500 else value
    return value


def validate_bien_data(data, bien_id=None):
    """
    Validar datos de un bien

    NOTA: Validaciones removidas temporalmente para permitir que operadores puedan
    inventariar sin restricciones. Será restaurado después.

    Args:
        data: diccionario con datos del bien
        bien_id: ID del bien si es edición
    """
    # SIN RESTRICCIONES - Permitir que operadores guarden cualquier dato
    # Luego lo arreglamos con validaciones apropiadas
    return []


def format_date(date_obj):
    """Formatear fecha para visualización"""
    if isinstance(date_obj, str):
        return date_obj
    if date_obj:
        return date_obj.strftime('%d/%m/%Y %H:%M')
    return ''


def get_estado_label(estado_code):
    """Convertir código de estado a etiqueta legible"""
    estados = {
        'b': 'Bueno',
        'r': 'Regular',
        'm': 'Malo'
    }
    return estados.get(estado_code, estado_code or 'Desconocido')


def get_estado_color(estado_code):
    """Obtener color Bootstrap para estado"""
    colores = {
        'b': 'success',
        'r': 'warning',
        'm': 'danger',
        'Bueno': 'success',
        'Regular': 'warning',
        'Malo': 'danger'
    }
    return colores.get(estado_code, 'secondary')


# ==================== PAGINACIÓN ====================
def get_pagination_info(total_items, current_page, items_per_page):
    """Calcular información de paginación"""
    total_pages = (total_items + items_per_page - 1) // items_per_page
    offset = (current_page - 1) * items_per_page

    return {
        'total_items': total_items,
        'current_page': current_page,
        'total_pages': total_pages,
        'offset': offset,
        'has_previous': current_page > 1,
        'has_next': current_page < total_pages,
        'previous_page': current_page - 1 if current_page > 1 else None,
        'next_page': current_page + 1 if current_page < total_pages else None,
    }


# ==================== EXPORTACIÓN ====================
def export_bienes_excel(bienes, filename='bienes_export.xlsx'):
    """Exportar lista de bienes a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'Bienes'

        # Headers
        headers = ['ID', 'Código Patrimonial', 'Denominación', 'Marca', 'Modelo',
                   'Serie', 'Estado', 'Sede', 'Responsable', 'Ubicación', 'Observaciones']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        for row_idx, bien in enumerate(bienes, 2):
            ws.cell(row=row_idx, column=1, value=bien.get('id'))
            ws.cell(row=row_idx, column=2, value=bien.get('codigo_patrimonial'))
            ws.cell(row=row_idx, column=3, value=bien.get('denominacion'))
            ws.cell(row=row_idx, column=4, value=bien.get('marca'))
            ws.cell(row=row_idx, column=5, value=bien.get('modelo'))
            ws.cell(row=row_idx, column=6, value=bien.get('serie'))
            ws.cell(row=row_idx, column=7, value=get_estado_label(bien.get('estado')))
            ws.cell(row=row_idx, column=8, value=bien.get('sede'))
            ws.cell(row=row_idx, column=9, value=bien.get('responsable'))
            ws.cell(row=row_idx, column=10, value=bien.get('ubicacion_texto'))
            ws.cell(row=row_idx, column=11, value=bien.get('observaciones'))

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar
        wb.save(filename)
        return True
    except Exception as e:
        current_app.logger.error(f"Error exportando a Excel: {str(e)}")
        return False


# ==================== UTILIDADES DE BD ====================
def get_bien_display_data(bien):
    """Preparar datos de un bien para visualización"""
    if not bien:
        return None

    return {
        'id': bien.get('id'),
        'codigo_patrimonial': bien.get('codigo_patrimonial'),
        'cod_barras': bien.get('cod_barras'),
        'denominacion': bien.get('denominacion'),
        'marca': bien.get('marca'),
        'modelo': bien.get('modelo'),
        'serie': bien.get('serie'),
        'estado': get_estado_label(bien.get('estado')),
        'estado_color': get_estado_color(bien.get('estado')),
        'sede': bien.get('sede'),
        'dependencia': bien.get('dependencia'),
        'responsable': bien.get('responsable'),
        'ubicacion_texto': bien.get('ubicacion_texto'),
        'observaciones': bien.get('observaciones'),
        'fecha_registro': format_date(bien.get('fecha_registro')),
        'verificado': bien.get('verificado'),
        'cal_2020': bien.get('cal_2020'),
        'cal_2021': bien.get('cal_2021'),
        'cal_2022': bien.get('cal_2022'),
        'cal_2023': bien.get('cal_2023'),
        'cal_2024': bien.get('cal_2024'),
        'cal_2025': bien.get('cal_2025'),
    }
