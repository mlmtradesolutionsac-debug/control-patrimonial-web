#!/usr/bin/env python3
"""
IMPORTAR archivo Excel SUPE CAL 2025 a PostgreSQL
Agregar los 76 bienes de SUPE del archivo DATA SIGA 03-12-2025 ARREGLADO.xlsx
"""

import sys
import os
from openpyxl import load_workbook
from datetime import datetime

sys.path.insert(0, '/home/patrimonio/control_patrimonial')
os.chdir('/home/patrimonio/control_patrimonial')

from wsgi import app
from app.models_sqlalchemy import Bien, Sede, Unidad
from app import db
from sqlalchemy import func

print("\n" + "="*90)
print("IMPORTACIÓN DE EXCEL: SUPE CAL 2025")
print("="*90)

excel_path = '/home/patrimonio/uploads/excel/DATA SIGA 03-12-2025 ARREGLADO.xlsx'

if not os.path.exists(excel_path):
    print(f"ERROR: Archivo no encontrado: {excel_path}")
    sys.exit(1)

print(f"\n✓ Archivo: {excel_path}")
print(f"  Tamaño: {os.path.getsize(excel_path) / 1024:.1f} KB")

# PASO 1: VERIFICAR ESTADO ANTES
print("\n" + "="*90)
print("PASO 1: ESTADO INICIAL")
print("="*90)

with app.app_context():
    total_antes = Bien.query.count()
    cal_antes = Bien.query.filter(Bien.cal_2025.isnot(None), Bien.cal_2025 != '').count()

    sede_supe = Sede.query.filter(Sede.nombre.ilike('SUPE')).first()
    supe_antes = Bien.query.filter_by(sede_id=sede_supe.id).count() if sede_supe else 0
    supe_cal_antes = Bien.query.filter_by(sede_id=sede_supe.id).filter(
        Bien.cal_2025.isnot(None), Bien.cal_2025 != ''
    ).count() if sede_supe else 0

    print(f"Total bienes: {total_antes:,}")
    print(f"Con CAL 2025: {cal_antes:,} ({(cal_antes/total_antes*100):.2f}%)")
    print(f"SUPE total: {supe_antes}")
    print(f"SUPE con CAL 2025: {supe_cal_antes}")

# PASO 2: CARGAR Y PROCESAR EXCEL
print("\n" + "="*90)
print("PASO 2: PROCESANDO ARCHIVO EXCEL")
print("="*90)

wb = load_workbook(excel_path)
ws = wb['DATOS ']

headers = [cell.value for cell in ws[1]]
print(f"Encabezados: {len(headers)} columnas")

# Encontrar índices de columnas
sede_idx = headers.index('Sede') + 1
cal_2025_idx = headers.index('CAL 2025') + 1
codigo_idx = headers.index('Codigo Patrimonial') + 1
denom_idx = headers.index('Denominacion') + 1
marca_idx = headers.index('Marca') + 1
serie_idx = headers.index('Serie') + 1
modelo_idx = headers.index('modelo') + 1
responsable_idx = headers.index('Responsable') + 1
dependencia_idx = headers.index('Dependencia') + 1
cal_2020_idx = headers.index('CAL 2020') + 1
cal_2021_idx = headers.index('CAL 2021') + 1
cal_2022_idx = headers.index('CAL 2022') + 1
cal_2023_idx = headers.index('CAL 2023') + 1
cal_2024_idx = headers.index('CAL 2024') + 1

# Procesar filas
filas_procesadas = 0
supe_encontrados = 0
supe_importados = 0
supe_actualizados = 0

print("\nProcesando filas...")
print("-" * 90)

bienes_a_insertar = []
bienes_a_actualizar = []

for row_idx in range(2, ws.max_row + 1):
    sede_val = ws.cell(row_idx, sede_idx).value
    cal_2025_val = ws.cell(row_idx, cal_2025_idx).value

    # Solo procesar SUPE con CAL 2025
    if not (sede_val and 'SUPE' in str(sede_val).upper() and cal_2025_val and str(cal_2025_val).strip() != ''):
        continue

    supe_encontrados += 1

    # Extraer datos
    codigo = ws.cell(row_idx, codigo_idx).value
    denom = ws.cell(row_idx, denom_idx).value
    marca = ws.cell(row_idx, marca_idx).value
    serie = ws.cell(row_idx, serie_idx).value
    modelo = ws.cell(row_idx, modelo_idx).value
    responsable = ws.cell(row_idx, responsable_idx).value
    dependencia = ws.cell(row_idx, dependencia_idx).value
    cal_2020 = ws.cell(row_idx, cal_2020_idx).value
    cal_2021 = ws.cell(row_idx, cal_2021_idx).value
    cal_2022 = ws.cell(row_idx, cal_2022_idx).value
    cal_2023 = ws.cell(row_idx, cal_2023_idx).value
    cal_2024 = ws.cell(row_idx, cal_2024_idx).value

    # Truncar campos CAL para evitar varchar overflow
    FIELD_LIMITS = {
        'cal_2020': 50,
        'cal_2021': 50,
        'cal_2022': 50,
        'cal_2023': 50,
        'cal_2024': 50,
        'cal_2025': 50,
    }

    cal_2020 = str(cal_2020)[:FIELD_LIMITS['cal_2020']] if cal_2020 else None
    cal_2021 = str(cal_2021)[:FIELD_LIMITS['cal_2021']] if cal_2021 else None
    cal_2022 = str(cal_2022)[:FIELD_LIMITS['cal_2022']] if cal_2022 else None
    cal_2023 = str(cal_2023)[:FIELD_LIMITS['cal_2023']] if cal_2023 else None
    cal_2024 = str(cal_2024)[:FIELD_LIMITS['cal_2024']] if cal_2024 else None
    cal_2025 = str(cal_2025_val)[:FIELD_LIMITS['cal_2025']] if cal_2025_val else None

    bienes_a_insertar.append({
        'codigo_patrimonial': codigo,
        'denominacion': denom,
        'marca': marca,
        'serie': serie,
        'modelo': modelo,
        'responsable': responsable,
        'dependencia': dependencia,
        'sede_id': 5,  # SUPE
        'cal_2020': cal_2020,
        'cal_2021': cal_2021,
        'cal_2022': cal_2022,
        'cal_2023': cal_2023,
        'cal_2024': cal_2024,
        'cal_2025': cal_2025,
    })

print(f"Bienes SUPE en Excel: {supe_encontrados}")
print(f"Bienes a procesar: {len(bienes_a_insertar)}")

# PASO 3: IMPORTAR A POSTGRESQL
print("\n" + "="*90)
print("PASO 3: IMPORTANDO A POSTGRESQL")
print("="*90)

with app.app_context():
    for idx, bien_data in enumerate(bienes_a_insertar, 1):
        try:
            # Verificar si ya existe por código
            existing = Bien.query.filter_by(codigo_patrimonial=bien_data['codigo_patrimonial']).first()

            if existing:
                # Actualizar
                existing.cal_2025 = bien_data['cal_2025']
                existing.cal_2024 = bien_data['cal_2024']
                existing.cal_2023 = bien_data['cal_2023']
                existing.cal_2022 = bien_data['cal_2022']
                existing.cal_2021 = bien_data['cal_2021']
                existing.cal_2020 = bien_data['cal_2020']
                supe_actualizados += 1
            else:
                # Insertar
                bien = Bien(
                    codigo_patrimonial=bien_data['codigo_patrimonial'],
                    denominacion=bien_data['denominacion'],
                    marca=bien_data['marca'],
                    serie=bien_data['serie'],
                    modelo=bien_data['modelo'],
                    responsable=bien_data['responsable'],
                    dependencia=bien_data['dependencia'],
                    sede_id=bien_data['sede_id'],
                    cal_2020=bien_data['cal_2020'],
                    cal_2021=bien_data['cal_2021'],
                    cal_2022=bien_data['cal_2022'],
                    cal_2023=bien_data['cal_2023'],
                    cal_2024=bien_data['cal_2024'],
                    cal_2025=bien_data['cal_2025'],
                    fecha_registro=datetime.now()
                )
                db.session.add(bien)
                supe_importados += 1

            # Commit cada 20 registros
            if idx % 20 == 0:
                db.session.commit()
                print(f"  ✓ {idx}/{len(bienes_a_insertar)} registros procesados")

        except Exception as e:
            db.session.rollback()
            print(f"  ✗ Error en fila {idx}: {e}")

    # Commit final
    db.session.commit()
    print(f"  ✓ {len(bienes_a_insertar)} registros procesados (FINAL)")

    # PASO 4: VERIFICAR ESTADO DESPUÉS
    print("\n" + "="*90)
    print("PASO 4: ESTADO FINAL")
    print("="*90)

    total_despues = Bien.query.count()
    cal_despues = Bien.query.filter(Bien.cal_2025.isnot(None), Bien.cal_2025 != '').count()

    sede_supe = Sede.query.filter(Sede.nombre.ilike('SUPE')).first()
    supe_despues = Bien.query.filter_by(sede_id=sede_supe.id).count() if sede_supe else 0
    supe_cal_despues = Bien.query.filter_by(sede_id=sede_supe.id).filter(
        Bien.cal_2025.isnot(None), Bien.cal_2025 != ''
    ).count() if sede_supe else 0

    print(f"ANTES:  {total_antes:,} bienes, {cal_antes:,} con CAL 2025")
    print(f"AHORA:  {total_despues:,} bienes, {cal_despues:,} con CAL 2025")
    print(f"CAMBIO: +{total_despues - total_antes:,} bienes, +{cal_despues - cal_antes:,} CAL 2025")

    print(f"\nSUPE ANTES:  {supe_antes} bienes, {supe_cal_antes} con CAL 2025")
    print(f"SUPE AHORA:  {supe_despues} bienes, {supe_cal_despues} con CAL 2025")
    print(f"SUPE CAMBIO: +{supe_cal_despues - supe_cal_antes:,} CAL 2025")

    avance_antes = (cal_antes / total_antes * 100) if total_antes > 0 else 0
    avance_despues = (cal_despues / total_despues * 100) if total_despues > 0 else 0

    print(f"\nAvance ANTES: {avance_antes:.2f}%")
    print(f"Avance AHORA: {avance_despues:.2f}%")
    print(f"Avance MEJORÓ: +{avance_despues - avance_antes:.2f}%")

    print(f"\n✓ IMPORTADOS: {supe_importados} nuevos bienes")
    print(f"✓ ACTUALIZADOS: {supe_actualizados} bienes existentes")

    # Validación final
    if total_despues >= 12700 and cal_despues >= 2400:
        print("\n✓✓✓ SISTEMA ÍNTEGRO - Importación exitosa ✓✓✓")
    else:
        print("\n✗ ALERTA: Los números bajaron, necesita investigación")

print("\n" + "="*90 + "\n")
