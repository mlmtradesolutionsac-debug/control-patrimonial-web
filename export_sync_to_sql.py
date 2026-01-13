#!/usr/bin/env python3
"""
Exporta los datos del Excel a formato SQL que puede ejecutarse en la BD PostgreSQL en la nube
"""

import sqlite3
from openpyxl import load_workbook
import json
from datetime import datetime

# Rutas
EXCEL_PATH = r"E:\PROYECTOS WEB\control_patrimonial_web\excel\DATA SIGA 15-12-2025.xlsx"
DB_PATH = r"E:\PROYECTOS WEB\control_patrimonial_web\data\inventario_patrimonial.db"
OUTPUT_SQL = r"E:\PROYECTOS WEB\control_patrimonial_web\sync_to_cloud.sql"

def map_estado(estado_str):
    """Convierte estado de texto a código (b=Bueno, r=Regular, m=Malo)"""
    if not estado_str:
        return None
    estado = estado_str.strip().upper()
    mapping = {
        'BUENO': 'b',
        'B': 'b',
        'REGULAR': 'r',
        'R': 'r',
        'MALO': 'm',
        'M': 'm'
    }
    return mapping.get(estado, None)

def escape_sql(text):
    """Escapa caracteres especiales para SQL"""
    if text is None:
        return 'NULL'
    if isinstance(text, str):
        return f"'{text.replace(chr(39), chr(39)*2)}'"
    return str(text)

def main():
    print("="*80)
    print("GENERANDO SQL PARA SINCRONIZACION EN CLOUD")
    print("="*80)

    # Conectar a BD local
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Cargar Excel
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Leer datos del Excel
    excel_bienes = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        if row[1]:  # si hay código patrimonial
            codigo = row[1]
            bien = {
                'Codigo Patrimonial': row[1],
                'Codigo de barras': row[2],
                'Denominacion': row[3],
                'Marca': row[4],
                'Serie': row[5],
                'modelo': row[6],
                'Responsable': row[7],
                'Usuario': row[8],
                'Sede': row[9],
                'Dependencia': row[10],
                'Estado': row[11],
                'observaciones': row[12],
                'CAL 2025': row[21]
            }
            excel_bienes[codigo] = bien

    print(f"Total bienes en Excel: {len(excel_bienes)}")

    # Obtener códigos de BD local
    cursor.execute("SELECT codigo_patrimonial, cal_2025 FROM bienes")
    db_bienes = {row[0]: row[1] for row in cursor.fetchall()}
    print(f"Total bienes en BD local: {len(db_bienes)}")

    # Obtener IDs de sedes y unidades
    cursor.execute("SELECT id, nombre FROM sedes")
    sedes = {row[1]: row[0] for row in cursor.fetchall()}

    cursor.execute("SELECT id, nombre FROM unidades")
    unidades = {row[1]: row[0] for row in cursor.fetchall()}

    # Identificar bienes nuevos
    bienes_nuevos = [cod for cod in excel_bienes if cod not in db_bienes]
    print(f"Bienes nuevos a insertar: {len(bienes_nuevos)}")

    # Identificar bienes a actualizar
    bienes_actualizar = []
    for cod, bien_excel in excel_bienes.items():
        if cod in db_bienes:
            cal_excel = bien_excel['CAL 2025']
            cal_db = db_bienes[cod]
            if cal_excel and cal_db != cal_excel:
                bienes_actualizar.append((cod, cal_excel))

    print(f"Bienes a actualizar CAL 2025: {len(bienes_actualizar)}")

    # Generar archivo SQL
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write("-- ========================================\n")
        f.write("-- SINCRONIZACION DE DATOS DEL EXCEL\n")
        f.write(f"-- Fecha: {datetime.now().isoformat()}\n")
        f.write("-- ========================================\n\n")

        # INSERT para bienes nuevos
        if bienes_nuevos:
            f.write("-- BIENES NUEVOS A INSERTAR\n")
            f.write(f"-- Total: {len(bienes_nuevos)}\n\n")

            for codigo in bienes_nuevos:
                bien = excel_bienes[codigo]

                codigo_patrimonial = bien['Codigo Patrimonial']
                cod_barras = bien['Codigo de barras']
                denominacion = escape_sql(bien['Denominacion'])
                marca = escape_sql(bien['Marca'])
                serie = escape_sql(bien['Serie'])
                modelo = escape_sql(bien['modelo'])
                responsable = escape_sql(bien['Responsable'])
                usuario_registro = escape_sql(bien['Usuario'] or 'ADMIN_SYNC')
                sede_nombre = bien['Sede']
                sede_id = sedes.get(sede_nombre, 'NULL')
                unidad_nombre = bien['Dependencia']
                unidad_id = unidades.get(unidad_nombre, 'NULL')
                estado = escape_sql(map_estado(bien['Estado']))
                observaciones = escape_sql(bien['observaciones'])
                cal_2025 = escape_sql(bien['CAL 2025'])
                verificado = 1

                f.write(f"INSERT INTO bienes (\n")
                f.write(f"    codigo_patrimonial, cod_barras, denominacion, marca, serie, modelo,\n")
                f.write(f"    responsable, usuario_registro, sede_id, unidad_id, estado,\n")
                f.write(f"    observaciones, verificado, cal_2025, fecha_registro\n")
                f.write(f") VALUES (\n")
                f.write(f"    {escape_sql(codigo_patrimonial)}, {escape_sql(cod_barras)}, {denominacion}, {marca}, {serie}, {modelo},\n")
                f.write(f"    {responsable}, {usuario_registro}, {sede_id}, {unidad_id}, {estado},\n")
                f.write(f"    {observaciones}, {verificado}, {cal_2025}, NOW()\n")
                f.write(f");\n\n")

        # UPDATE para CAL 2025
        if bienes_actualizar:
            f.write("-- ACTUALIZACION DE CAL 2025\n")
            f.write(f"-- Total: {len(bienes_actualizar)}\n\n")

            for codigo, cal_2025 in bienes_actualizar:
                f.write(f"UPDATE bienes SET cal_2025 = {escape_sql(cal_2025)} ")
                f.write(f"WHERE codigo_patrimonial = {escape_sql(codigo)};\n")

        f.write("\n-- ========================================\n")
        f.write("-- FIN DE SINCRONIZACION\n")
        f.write("-- ========================================\n")

    print(f"\n[OK] Archivo SQL generado: {OUTPUT_SQL}")
    print(f"Tamaño: {len(open(OUTPUT_SQL).read())} bytes")

    cursor.close()
    conn.close()

if __name__ == '__main__':
    main()
