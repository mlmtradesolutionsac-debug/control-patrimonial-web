#!/usr/bin/env python3
"""
Script para sincronizar datos del Excel a la base de datos
- Inserta bienes nuevos que no están en la BD
- Actualiza CAL 2025 para bienes existentes
"""

import sqlite3
from datetime import datetime
from openpyxl import load_workbook
import json

# Rutas
EXCEL_PATH = r"E:\PROYECTOS WEB\control_patrimonial_web\excel\DATA SIGA 15-12-2025.xlsx"
DB_PATH = r"E:\PROYECTOS WEB\control_patrimonial_web\data\inventario_patrimonial.db"

def map_estado(estado_str):
    """Convierte estado de texto a código (b=Bueno, r=Regular, m=Malo)"""
    if not estado_str:
        return None
    estado = estado_str.strip().upper()
    mapping = {
        'BUENO': 'b',
        'B': 'b',
        'REGULAR': 'r',
        'R': 'r',
        'MALO': 'm',
        'M': 'm'
    }
    return mapping.get(estado, None)

def get_sede_id(conn, sede_nombre):
    """Obtiene el ID de sede por nombre"""
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM sedes WHERE nombre = ?", (sede_nombre,))
    result = cursor.fetchone()
    return result[0] if result else None

def get_unidad_id(conn, unidad_nombre):
    """Obtiene el ID de unidad por nombre"""
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM unidades WHERE nombre = ?", (unidad_nombre,))
    result = cursor.fetchone()
    return result[0] if result else None

def insert_bien(conn, bien_data, usuario):
    """Inserta un bien nuevo en la BD"""
    cursor = conn.cursor()

    try:
        # Preparar datos
        codigo_patrimonial = bien_data['Codigo Patrimonial']
        cod_barras = bien_data['Codigo de barras']
        denominacion = bien_data['Denominacion']
        marca = bien_data.get('Marca')
        serie = bien_data.get('Serie')
        modelo = bien_data.get('modelo')
        responsable = bien_data.get('Responsable')
        usuario_registro = bien_data.get('Usuario', usuario)
        sede_nombre = bien_data.get('Sede')
        unidad_nombre = bien_data.get('Dependencia')
        estado_str = bien_data.get('Estado')
        estado = map_estado(estado_str)
        observaciones = bien_data.get('observaciones')
        cal_2025 = bien_data.get('CAL 2025')

        # Obtener IDs
        sede_id = get_sede_id(conn, sede_nombre)
        unidad_id = get_unidad_id(conn, unidad_nombre)

        if not sede_id:
            print(f"  ERROR: No se encontró sede '{sede_nombre}'")
            return False

        # Insertar bien
        cursor.execute("""
            INSERT INTO bienes (
                codigo_patrimonial, cod_barras, denominacion, marca, serie, modelo,
                responsable, usuario_registro, sede_id, unidad_id, estado,
                observaciones, verificado, cal_2025, fecha_registro
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            codigo_patrimonial, cod_barras, denominacion, marca, serie, modelo,
            responsable, usuario_registro, sede_id, unidad_id, estado,
            observaciones, 1, cal_2025, datetime.now().isoformat()
        ))

        bien_id = cursor.lastrowid

        # Registrar en historial
        detalle = json.dumps({
            'codigo_patrimonial': codigo_patrimonial,
            'denominacion': denominacion,
            'cal_2025': cal_2025
        })

        cursor.execute("""
            INSERT INTO historial (bien_id, usuario, accion, detalle_json, fecha)
            VALUES (?, ?, ?, ?, ?)
        """, (bien_id, usuario_registro, 'CREAR', detalle, datetime.now().isoformat()))

        conn.commit()
        print(f"  [OK] Bien insertado: {codigo_patrimonial} (ID: {bien_id})")
        return True

    except sqlite3.IntegrityError as e:
        print(f"  ERROR: {e}")
        conn.rollback()
        return False
    except Exception as e:
        print(f"  ERROR: {e}")
        conn.rollback()
        return False

def update_cal_2025(conn, codigos_a_actualizar, usuario):
    """Actualiza CAL 2025 para múltiples bienes"""
    cursor = conn.cursor()
    actualizados = 0
    errores = 0

    for codigo, nuevo_cal in codigos_a_actualizar:
        try:
            # Obtener bien actual
            cursor.execute(
                "SELECT id, cal_2025 FROM bienes WHERE codigo_patrimonial = ?",
                (codigo,)
            )
            result = cursor.fetchone()

            if not result:
                print(f"  ERROR: Bien {codigo} no encontrado en BD")
                errores += 1
                continue

            bien_id, cal_actual = result

            # Actualizar
            cursor.execute(
                "UPDATE bienes SET cal_2025 = ? WHERE id = ?",
                (nuevo_cal, bien_id)
            )

            # Registrar en historial
            detalle = json.dumps({
                'campo': 'cal_2025',
                'anterior': cal_actual,
                'nuevo': nuevo_cal
            })

            cursor.execute("""
                INSERT INTO historial (bien_id, usuario, accion, detalle_json, fecha)
                VALUES (?, ?, ?, ?, ?)
            """, (bien_id, usuario, 'EDITAR', detalle, datetime.now().isoformat()))

            actualizados += 1

        except Exception as e:
            print(f"  ERROR al actualizar {codigo}: {e}")
            errores += 1

    conn.commit()
    return actualizados, errores

def main():
    print("="*80)
    print("SINCRONIZACION EXCEL -> DATABASE")
    print("="*80)

    # Conectar a BD
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Cargar Excel
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Obtener encabezados
    headers = [cell.value for cell in ws[1]]

    # Leer datos del Excel
    excel_bienes = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        if row[1]:  # si hay código patrimonial
            codigo = row[1]
            bien = {
                'N': row[0],
                'Codigo Patrimonial': row[1],
                'Codigo de barras': row[2],
                'Denominacion': row[3],
                'Marca': row[4],
                'Serie': row[5],
                'modelo': row[6],
                'Responsable': row[7],
                'Usuario': row[8],
                'Sede': row[9],
                'Dependencia': row[10],
                'Estado': row[11],
                'observaciones': row[12],
                'Verificado': row[13],
                'SIGA TIPO': row[14],
                'Siga UBICACION': row[15],
                'CAL 2025': row[21]
            }
            excel_bienes[codigo] = bien

    print(f"\nTotal bienes en Excel: {len(excel_bienes)}")

    # Obtener códigos patrimoniales de BD
    cursor.execute("SELECT codigo_patrimonial, cal_2025 FROM bienes")
    db_bienes = {row[0]: row[1] for row in cursor.fetchall()}
    print(f"Total bienes en DB: {len(db_bienes)}\n")

    # Identificar bienes nuevos
    bienes_nuevos = [cod for cod in excel_bienes if cod not in db_bienes]
    print(f"BIENES NUEVOS A INSERTAR: {len(bienes_nuevos)}")
    print("="*80)

    if bienes_nuevos:
        for codigo in bienes_nuevos:
            bien = excel_bienes[codigo]
            print(f"Insertando: {codigo} - {bien['Denominacion'][:50]}")
            insert_bien(conn, bien, "ADMIN_SYNC")

    # Identificar bienes a actualizar CAL 2025
    bienes_actualizar = []
    for cod, bien_excel in excel_bienes.items():
        if cod in db_bienes:
            cal_excel = bien_excel['CAL 2025']
            cal_db = db_bienes[cod]
            if cal_excel and cal_db != cal_excel:
                bienes_actualizar.append((cod, cal_excel))

    print(f"\nBIENES A ACTUALIZAR CAL 2025: {len(bienes_actualizar)}")
    print("="*80)

    if bienes_actualizar:
        print("Actualizando CAL 2025...")
        actualizados, errores = update_cal_2025(conn, bienes_actualizar, "ADMIN_SYNC")
        print(f"[OK] Actualizados: {actualizados}")
        if errores:
            print(f"[ERROR] Errores: {errores}")

    # Resumen final
    print("\n" + "="*80)
    print("RESUMEN")
    print("="*80)
    cursor.execute("SELECT COUNT(*) FROM bienes")
    total_bienes = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM bienes WHERE cal_2025 IS NOT NULL")
    bienes_con_cal = cursor.fetchone()[0]

    print(f"Total bienes en BD: {total_bienes}")
    print(f"Bienes con CAL 2025: {bienes_con_cal}")
    print(f"Bienes sin CAL 2025: {total_bienes - bienes_con_cal}")

    conn.close()
    print("\n[OK] Sincronizacion completada")

if __name__ == '__main__':
    main()
