#!/usr/bin/env python3
"""
Analizar archivo Excel para ver cuántos bienes SUPE con CAL 2025 contiene
"""

import sys
import os
from openpyxl import load_workbook

print("\n" + "="*90)
print("ANÁLISIS DE ARCHIVO EXCEL - SUPE CAL 2025")
print("="*90)

excel_path = '/home/patrimonio/uploads/excel/DATA SIGA 03-12-2025 ARREGLADO.xlsx'

if not os.path.exists(excel_path):
    print(f"ERROR: Archivo no encontrado: {excel_path}")
    sys.exit(1)

try:
    # Cargar el workbook
    wb = load_workbook(excel_path)
    print(f"\n✓ Archivo cargado: {excel_path}")
    print(f"  Tamaño: {os.path.getsize(excel_path) / 1024:.1f} KB")
    print(f"  Hojas: {wb.sheetnames}")

    # Analizar cada hoja
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*90}")
        print(f"HOJA: {sheet_name}")
        print(f"{'='*90}")

        ws = wb[sheet_name]

        # Obtener encabezados
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        print(f"\nEncabezados ({len(headers)} columnas):")
        print("-" * 90)
        for i, header in enumerate(headers, 1):
            print(f"  {i:2}. {header}")

        # Contar filas de datos
        data_rows = ws.max_row - 1  # Excluir encabezado
        print(f"\nTotal filas de datos: {data_rows}")

        # Buscar columnas relevantes
        sede_col = None
        cal_2025_col = None
        codigo_col = None

        for i, header in enumerate(headers, 1):
            if header and 'sede' in str(header).lower():
                sede_col = i
            if header and ('cal 2025' in str(header).lower() or 'cal_2025' in str(header).lower()):
                cal_2025_col = i
            if header and 'código' in str(header).lower():
                codigo_col = i

        print(f"\nColumnas detectadas:")
        print(f"  Sede: Columna {sede_col}")
        print(f"  CAL 2025: Columna {cal_2025_col}")
        print(f"  Código: Columna {codigo_col}")

        # Contar bienes SUPE con CAL 2025
        bienes_supe = 0
        supe_con_cal = 0
        supe_sin_cal = 0

        print(f"\nAnalizando {data_rows} filas...")
        print("-" * 90)

        for row_idx in range(2, ws.max_row + 1):
            # Obtener valores
            sede_val = ws.cell(row_idx, sede_col).value if sede_col else None
            cal_val = ws.cell(row_idx, cal_2025_col).value if cal_2025_col else None
            codigo_val = ws.cell(row_idx, codigo_col).value if codigo_col else None

            # Buscar SUPE
            if sede_val and 'SUPE' in str(sede_val).upper():
                bienes_supe += 1
                if cal_val and str(cal_val).strip() != '':
                    supe_con_cal += 1
                else:
                    supe_sin_cal += 1

        print(f"  Bienes SUPE encontrados: {bienes_supe}")
        print(f"  SUPE con CAL 2025: {supe_con_cal}")
        print(f"  SUPE sin CAL 2025: {supe_sin_cal}")

        # Mostrar primeros 10 registros SUPE con CAL 2025
        if supe_con_cal > 0:
            print(f"\nPrimeros 10 bienes SUPE CON CAL 2025:")
            print("-" * 90)
            print(f"{'Código':<20} {'Sede':<30} {'CAL 2025':<20}")
            print("-" * 90)

            count = 0
            for row_idx in range(2, ws.max_row + 1):
                sede_val = ws.cell(row_idx, sede_col).value if sede_col else None
                cal_val = ws.cell(row_idx, cal_2025_col).value if cal_2025_col else None
                codigo_val = ws.cell(row_idx, codigo_col).value if codigo_col else None

                if sede_val and 'SUPE' in str(sede_val).upper() and cal_val and str(cal_val).strip() != '':
                    print(f"{str(codigo_val):<20} {str(sede_val):<30} {str(cal_val):<20}")
                    count += 1
                    if count >= 10:
                        break

    # RESUMEN FINAL
    print(f"\n{'='*90}")
    print("RESUMEN:")
    print(f"{'='*90}")

    total_supe_en_excel = 0
    total_cal_2025_en_excel = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        sede_col = None
        cal_2025_col = None

        for i, header in enumerate(headers, 1):
            if header and 'sede' in str(header).lower():
                sede_col = i
            if header and ('cal 2025' in str(header).lower() or 'cal_2025' in str(header).lower()):
                cal_2025_col = i

        if sede_col and cal_2025_col:
            for row_idx in range(2, ws.max_row + 1):
                sede_val = ws.cell(row_idx, sede_col).value
                cal_val = ws.cell(row_idx, cal_2025_col).value

                if sede_val and 'SUPE' in str(sede_val).upper():
                    total_supe_en_excel += 1
                    if cal_val and str(cal_val).strip() != '':
                        total_cal_2025_en_excel += 1

    print(f"\n✓ Total SUPE EN ARCHIVO EXCEL: {total_supe_en_excel}")
    print(f"✓ SUPE CON CAL 2025 EN ARCHIVO: {total_cal_2025_en_excel}")

    print(f"\n{'='*90}")
    print("CONCLUSIÓN:")
    print(f"{'='*90}")
    print(f"\nEste archivo contiene {total_cal_2025_en_excel} bienes de SUPE con CAL 2025")
    print(f"que DEBERÍAN estar en el sistema pero NO están.")
    print(f"\nNecesitamos IMPORTAR este archivo a PostgreSQL ahora.")

    print(f"\n{'='*90}\n")

except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
