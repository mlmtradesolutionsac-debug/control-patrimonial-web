#!/usr/bin/env python3
"""
Analizar E IMPORTAR segundo archivo Excel: DATA SIGA 03-12-2025 ultimo enita.xlsx
"""

import sys
import os
from openpyxl import load_workbook
from datetime import datetime

sys.path.insert(0, '/home/patrimonio/control_patrimonial')
os.chdir('/home/patrimonio/control_patrimonial')

from wsgi import app
from app.models_sqlalchemy import Bien, Sede, Unidad
from app import db

print("\n" + "="*90)
print("ANÁLISIS E IMPORTACIÓN: DATA SIGA 03-12-2025 ultimo enita.xlsx")
print("="*90)

excel_path = '/home/patrimonio/uploads/excel/DATA SIGA 03-12-2025 ultimo enita.xlsx'

if not os.path.exists(excel_path):
    print(f"ERROR: Archivo no encontrado: {excel_path}")
    sys.exit(1)

print(f"\n✓ Archivo: {excel_path}")
print(f"  Tamaño: {os.path.getsize(excel_path) / 1024:.1f} KB")

# PASO 1: ESTADO ANTES
print("\n" + "="*90)
print("PASO 1: ESTADO INICIAL")
print("="*90)

with app.app_context():
    total_antes = Bien.query.count()
    cal_antes = Bien.query.filter(Bien.cal_2025.isnot(None), Bien.cal_2025 != '').count()

    sede_supe = Sede.query.filter(Sede.nombre.ilike('SUPE')).first()
    supe_antes = Bien.query.filter_by(sede_id=sede_supe.id).count() if sede_supe else 0
    supe_cal_antes = Bien.query.filter_by(sede_id=sede_supe.id).filter(
        Bien.cal_2025.isnot(None), Bien.cal_2025 != ''
    ).count() if sede_supe else 0

    print(f"Total bienes: {total_antes:,}")
    print(f"Con CAL 2025: {cal_antes:,}")
    print(f"SUPE total: {supe_antes}")
    print(f"SUPE con CAL 2025: {supe_cal_antes}")

# PASO 2: ANALIZAR CONTENIDO
print("\n" + "="*90)
print("PASO 2: ANALIZANDO CONTENIDO DEL ARCHIVO")
print("="*90)

wb = load_workbook(excel_path)
print(f"Hojas disponibles: {wb.sheetnames}")

ws = wb[wb.sheetnames[0]]  # Primera hoja
headers = [cell.value for cell in ws[1]]

print(f"\nEncabezados: {len(headers)} columnas")
print(f"Total filas: {ws.max_row - 1}")

# Buscar columnas
sede_idx = None
cal_2025_idx = None
codigo_idx = None

for i, h in enumerate(headers, 1):
    if h and 'sede' in str(h).lower():
        sede_idx = i
    if h and ('cal 2025' in str(h).lower() or 'cal_2025' in str(h).lower()):
        cal_2025_idx = i
    if h and 'código' in str(h).lower():
        codigo_idx = i

print(f"Sede (columna {sede_idx}), CAL 2025 (columna {cal_2025_idx}), Código (columna {codigo_idx})")

# Contar SUPE
bienes_supe = 0
supe_con_cal = 0

for row_idx in range(2, ws.max_row + 1):
    sede_val = ws.cell(row_idx, sede_idx).value if sede_idx else None
    cal_val = ws.cell(row_idx, cal_2025_idx).value if cal_2025_idx else None

    if sede_val and 'SUPE' in str(sede_val).upper():
        bienes_supe += 1
        if cal_val and str(cal_val).strip() != '':
            supe_con_cal += 1

print(f"\n✓ Bienes SUPE en este archivo: {bienes_supe}")
print(f"✓ SUPE con CAL 2025: {supe_con_cal}")

# PASO 3: IMPORTAR
if supe_con_cal > 0:
    print("\n" + "="*90)
    print(f"PASO 3: IMPORTANDO {supe_con_cal} BIENES SUPE CON CAL 2025")
    print("="*90)

    # Recolectar datos
    todas_las_columnas = {}
    for i, h in enumerate(headers, 1):
        todas_las_columnas[h] = i

    bienes_a_procesar = []

    for row_idx in range(2, ws.max_row + 1):
        sede_val = ws.cell(row_idx, sede_idx).value if sede_idx else None
        cal_val = ws.cell(row_idx, cal_2025_idx).value if cal_2025_idx else None

        if not (sede_val and 'SUPE' in str(sede_val).upper() and cal_val and str(cal_val).strip() != ''):
            continue

        bien_data = {}
        for col_name, col_idx in todas_las_columnas.items():
            bien_data[col_name] = ws.cell(row_idx, col_idx).value

        bienes_a_procesar.append(bien_data)

    # Importar
    with app.app_context():
        importados = 0
        actualizados = 0

        for idx, bien_data in enumerate(bienes_a_procesar, 1):
            try:
                codigo = bien_data.get('Codigo Patrimonial')
                existing = Bien.query.filter_by(codigo_patrimonial=codigo).first()

                if existing:
                    # Actualizar CAL 2025
                    existing.cal_2025 = str(bien_data.get('CAL 2025', ''))[:50]
                    actualizados += 1
                else:
                    # Insertar nuevo
                    bien = Bien(
                        codigo_patrimonial=codigo,
                        denominacion=bien_data.get('Denominacion'),
                        marca=bien_data.get('Marca'),
                        serie=bien_data.get('Serie'),
                        modelo=bien_data.get('modelo'),
                        responsable=bien_data.get('Responsable'),
                        dependencia=bien_data.get('Dependencia'),
                        sede_id=5,  # SUPE
                        cal_2020=str(bien_data.get('CAL 2020', ''))[:50] if bien_data.get('CAL 2020') else None,
                        cal_2021=str(bien_data.get('CAL 2021', ''))[:50] if bien_data.get('CAL 2021') else None,
                        cal_2022=str(bien_data.get('CAL 2022', ''))[:50] if bien_data.get('CAL 2022') else None,
                        cal_2023=str(bien_data.get('CAL 2023', ''))[:50] if bien_data.get('CAL 2023') else None,
                        cal_2024=str(bien_data.get('CAL 2024', ''))[:50] if bien_data.get('CAL 2024') else None,
                        cal_2025=str(bien_data.get('CAL 2025', ''))[:50] if bien_data.get('CAL 2025') else None,
                        fecha_registro=datetime.now()
                    )
                    db.session.add(bien)
                    importados += 1

                if idx % 10 == 0:
                    db.session.commit()
                    print(f"  ✓ {idx}/{len(bienes_a_procesar)} procesados")

            except Exception as e:
                db.session.rollback()
                print(f"  ✗ Error fila {idx}: {e}")

        db.session.commit()
        print(f"  ✓ {len(bienes_a_procesar)} bienes procesados (FINAL)")

        # PASO 4: VERIFICAR ESTADO FINAL
        print("\n" + "="*90)
        print("PASO 4: ESTADO FINAL")
        print("="*90)

        total_despues = Bien.query.count()
        cal_despues = Bien.query.filter(Bien.cal_2025.isnot(None), Bien.cal_2025 != '').count()

        sede_supe = Sede.query.filter(Sede.nombre.ilike('SUPE')).first()
        supe_despues = Bien.query.filter_by(sede_id=sede_supe.id).count() if sede_supe else 0
        supe_cal_despues = Bien.query.filter_by(sede_id=sede_supe.id).filter(
            Bien.cal_2025.isnot(None), Bien.cal_2025 != ''
        ).count() if sede_supe else 0

        print(f"ANTES:  {total_antes:,} bienes, {cal_antes:,} con CAL 2025")
        print(f"AHORA:  {total_despues:,} bienes, {cal_despues:,} con CAL 2025")
        print(f"CAMBIO: +{total_despues - total_antes:,} bienes, +{cal_despues - cal_antes:,} CAL 2025")

        print(f"\nSUPE ANTES:  {supe_antes} bienes, {supe_cal_antes} con CAL 2025")
        print(f"SUPE AHORA:  {supe_despues} bienes, {supe_cal_despues} con CAL 2025")
        print(f"SUPE CAMBIO: +{supe_cal_despues - supe_cal_antes:,} CAL 2025")

        avance_antes = (cal_antes / total_antes * 100) if total_antes > 0 else 0
        avance_despues = (cal_despues / total_despues * 100) if total_despues > 0 else 0

        print(f"\nAvance ANTES: {avance_antes:.2f}%")
        print(f"Avance AHORA: {avance_despues:.2f}%")
        print(f"Avance MEJORÓ: +{avance_despues - avance_antes:.2f}%")

        print(f"\n✓ IMPORTADOS: {importados} nuevos bienes")
        print(f"✓ ACTUALIZADOS: {actualizados} bienes existentes")

        if total_despues >= 12700 and cal_despues >= 2400:
            print("\n✓✓✓ SISTEMA ÍNTEGRO - Importación exitosa ✓✓✓")
        else:
            print("\n✗ ALERTA: Valores por debajo de umbral")

else:
    print(f"⚠️ Este archivo NO tiene bienes SUPE con CAL 2025")
    print(f"Total SUPE encontrados: {bienes_supe}")
    print(f"Con CAL 2025: {supe_con_cal}")

print("\n" + "="*90 + "\n")
